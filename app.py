from flask import (
    Flask,
    request,
    redirect,
    url_for,
    render_template,
    jsonify,
    session,
    flash,
    get_flashed_messages,
    send_file,
)
import os
import uuid
import re
import locale
from datetime import datetime
from functools import wraps
from io import BytesIO

import pandas as pd
import xlrd  # Para ler arquivos .xls, se necessário
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook  # Usado para trabalhar com XLSX
from openpyxl.utils import get_column_letter  # Para obter a coluna em letra
from openpyxl.cell import MergedCell  # Para identificar células mescladas


# ==========================================================
#  CONFIGURAÇÃO BÁSICA DA APLICAÇÃO
# ==========================================================

# Tenta definir a localidade para formatação de datas em português
try:
    locale.setlocale(locale.LC_TIME, "pt_BR.UTF-8")
except locale.Error:
    pass

app = Flask(__name__)
app.secret_key = "sua_chave_secreta"  # Altere para uma chave segura
ACCESS_TOKEN = "minha_senha"  # Token de acesso

app.config["UPLOAD_FOLDER"] = "uploads"
ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".gif"}
app.config["HOLIDAYS_JSON_PATH"] = os.path.join(os.path.dirname(__file__), "modelos", "feriados_nacionais.json")


# Cria os diretórios necessários, se não existirem
os.makedirs("static/fotos", exist_ok=True)
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

# Caminho do arquivo CSV relativo ao diretório do script
CSV_PATH = os.path.join(os.path.dirname(__file__), "uploads", "escolas.csv")

# Variável global para armazenar os dados do CSV
escolas_df = None


# ==========================================================
#  FUNÇÕES AUXILIARES – ESCOLAS (escolas.csv)
# ==========================================================

def carregar_escolas():
    """Carrega o CSV de escolas em um DataFrame global."""
    global escolas_df
    if os.path.exists(CSV_PATH):
        try:
            escolas_df = pd.read_csv(CSV_PATH, encoding="latin1", sep=";")
            print(f"[INFO] Arquivo {CSV_PATH} carregado com sucesso.")
        except Exception as e:
            escolas_df = None
            print(f"[ERRO] Falha ao carregar {CSV_PATH}: {e}")
    else:
        escolas_df = None
        print(f"[ERRO] Arquivo {CSV_PATH} não encontrado.")


def get_escolas_df():
    """Garante que o DataFrame de escolas está carregado."""
    global escolas_df
    if escolas_df is None or escolas_df.empty:
        print("[INFO] Recarregando arquivo escolas.csv...")
        carregar_escolas()
    return escolas_df


@app.before_request
def inicializar_escolas():
    """Garante que escolas.csv está carregado antes de cada requisição."""
    global escolas_df
    if escolas_df is None or (isinstance(escolas_df, pd.DataFrame) and escolas_df.empty):
        carregar_escolas()


@app.route("/escolas/search")
def escolas_search():
    """Endpoint para o Select2 buscar escolas no CSV."""
    df = get_escolas_df()
    query = request.args.get("q", "").lower().strip()
    results = []

    if df is not None and not df.empty and query:
        # Filtra usando pandas (assumindo coluna 3 = nome da escola)
        df_filtered = df[df.iloc[:, 3].str.lower().str.contains(query, na=False)]

        # Limita a 50 resultados para não sobrecarregar
        df_filtered = df_filtered.head(50)

        for _, row in df_filtered.iterrows():
            nome = str(row[3]).strip()
            municipio = str(row[2]).strip()
            uf = str(row[1]).strip()
            text = f"{nome} - {municipio}/{uf}"
            results.append({"id": nome, "text": text})

    return jsonify(results)


# Carrega o CSV na inicialização do sistema
carregar_escolas()


# ==========================================================
#  BLUEPRINTS
# ==========================================================

from confere import confere_bp

app.register_blueprint(confere_bp, url_prefix="/confere")


# ==========================================================
#  HELPERS GERAIS
# ==========================================================

def allowed_file(filename):
    _, ext = os.path.splitext(filename)
    return ext.lower() in ALLOWED_EXTENSIONS


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login_route", next=request.url))
        return f(*args, **kwargs)

    return decorated_function


def set_merged_cell_value(ws, cell_coord, value):
    """
    Atualiza o valor de uma célula mesclada em uma planilha openpyxl
    preservando a mesclagem.
    """
    cell = ws[cell_coord]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell_coord in merged_range:
                range_str = str(merged_range)
                ws.unmerge_cells(range_str)
                min_col, min_row, _, _ = merged_range.bounds
                top_left_coord = f"{get_column_letter(min_col)}{min_row}"
                ws[top_left_coord] = value
                ws.merge_cells(range_str)
                return
    ws[cell_coord] = value


def convert_xls_to_xlsx(file_like):
    """
    Converte um arquivo XLS (file-like) para um Workbook do openpyxl.
    """
    book_xlrd = xlrd.open_workbook(file_contents=file_like.read())
    wb = Workbook()
    # Remove a planilha padrão criada pelo openpyxl, se houver
    if "Sheet" in wb.sheetnames and len(book_xlrd.sheet_names()) > 0:
        std = wb.active
        wb.remove(std)

    for sheet_name in book_xlrd.sheet_names():
        sheet_xlrd = book_xlrd.sheet_by_name(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        for row in range(sheet_xlrd.nrows):
            for col in range(sheet_xlrd.ncols):
                ws.cell(row=row + 1, column=col + 1, value=sheet_xlrd.cell_value(row, col))

    return wb


def load_workbook_model(file):
    """
    Abre o arquivo do modelo XLSX (ou XLS convertendo-o para XLSX)
    preservando a formatação.
    """
    ext = os.path.splitext(file.filename)[1].lower()
    file.seek(0)
    if ext == ".xlsx":
        return load_workbook(file, data_only=False)
    elif ext == ".xls":
        content = file.read()
        return convert_xls_to_xlsx(BytesIO(content))
    else:
        raise ValueError("Formato de arquivo não suportado para o quadro modelo.")


def data_extenso(dt):
    """Retorna a data por extenso em português."""
    meses = [
        "janeiro",
        "fevereiro",
        "março",
        "abril",
        "maio",
        "junho",
        "julho",
        "agosto",
        "setembro",
        "outubro",
        "novembro",
        "dezembro",
    ]
    return f"{dt.day} de {meses[dt.month - 1]} de {dt.year}"


def is_valid_plano(val):
    """Avalia se o valor do Plano de Ação é considerado válido."""
    if val is None:
        return False
    s = str(val).strip()
    return s not in ["", "-", "0", "#REF"]


# ==========================================================
#  CARTEIRINHAS – GERAÇÃO DE HTML
# ==========================================================

def gerar_html_carteirinhas(arquivo_excel, somente_com_foto=False):
    # Lê a planilha do Fundamental
    planilha = pd.read_excel(arquivo_excel, sheet_name="LISTA CORRIDA")

    dados = planilha[["RM", "NOME", "DATA NASC.", "RA", "SAI SOZINHO?", "SÉRIE", "HORÁRIO"]].copy()
    dados["RM"] = dados["RM"].fillna(0).astype(int)

    # Filtra apenas alunos com RM válido
    registros_validos = []
    for _, row in dados.iterrows():
        rm_str = str(row["RM"]).strip()
        if not rm_str or rm_str == "0":
            continue
        registros_validos.append(row)

    alunos_sem_fotos_list = []
    alunos = []

    allowed_exts = [".jpg", ".jpeg", ".png", ".bmp", ".gif"]

    for row in registros_validos:
        nome = row["NOME"]
        data_nasc = row["DATA NASC."]
        serie = row["SÉRIE"]
        horario = row["HORÁRIO"]

        # Data de nascimento
        if pd.notna(data_nasc):
            try:
                data_nasc = pd.to_datetime(data_nasc, errors="coerce")
                if pd.notna(data_nasc):
                    data_nasc_str = data_nasc.strftime("%d/%m/%Y")
                else:
                    data_nasc_str = "Desconhecida"
            except Exception:
                data_nasc_str = "Desconhecida"
        else:
            data_nasc_str = "Desconhecida"

        ra = row["RA"]

        # Sai sozinho?
        sai_sozinho_raw = str(row["SAI SOZINHO?"]).strip().upper()
        if sai_sozinho_raw in ("SIM", "S", "YES", "Y"):
            classe_cor = "verde"
            status_texto = "Sai sozinho"
            status_icon = "&#10003;"
        else:
            classe_cor = "vermelho"
            status_texto = "Não sai sozinho"
            status_icon = "&#9888;"

        # Foto
        foto_url = None
        for ext in allowed_exts:
            caminho_foto = f"static/fotos/{row['RM']}{ext}"
            if os.path.exists(caminho_foto):
                foto_url = f"/static/fotos/{row['RM']}{ext}"
                break

        if not foto_url:
            alunos_sem_fotos_list.append(
                {
                    "rm": int(row["RM"]),
                    "nome": nome,
                    "serie": serie,
                }
            )

        alunos.append(
            {
                "rm": int(row["RM"]),
                "nome": nome,
                "data_nasc": data_nasc_str,
                "ra": ra,
                "serie": serie,
                "horario": horario,
                "classe_cor": classe_cor,
                "status_texto": status_texto,
                "status_icon": status_icon,
                "foto_url": foto_url,
            }
        )

    # --- AQUI ENTRA O FILTRO ---
    if somente_com_foto:
        alunos_para_exibir = [a for a in alunos if a.get("foto_url")]
    else:
        alunos_para_exibir = alunos

    # Paginação: 6 carteirinhas por página
    pages = []
    for i in range(0, len(alunos_para_exibir), 6):
        pages.append(alunos_para_exibir[i: i + 6])

    total_sem_foto = len(alunos_sem_fotos_list)

    return render_template(
        "gerar_carteirinhas.html",
        pages=pages,
        alunos_sem_foto=alunos_sem_fotos_list,
        total_sem_foto=total_sem_foto,
        somente_com_foto=somente_com_foto,  # opcional: p/ mostrar estado do filtro na tela
    )



# ==========================================================
#  DECLARAÇÕES – GERAÇÃO HTML (SINGULAR)
# ==========================================================

def gerar_declaracao_escolar(
    file_path,
    rm,
    tipo,
    file_path2=None,
    deve_historico=False,
    unidade_anterior=None,
    dados_frequencia=None,
):
    """
    Gera o HTML de uma declaração escolar (Escolaridade, Transferência, Conclusão
    ou Frequência) tanto para Fundamental quanto EJA, de acordo com
    session['declaracao_tipo'].

    file_path  -> caminho/arquivo padrão da lista piloto (salvo em sessão/ao entrar no sistema)
    file_path2 -> caminho/arquivo opcional, usado quando o usuário reenviar a lista
                  (por exemplo, após o servidor free acordar). SE informado, TERÁ PRIORIDADE.
    dados_frequencia -> dicionário opcional com os dados de frequência por mês,
                        utilizado apenas quando tipo == "Frequencia".
    """
    global escolas_df

    # Se um segundo caminho foi informado (lista reenviada), ele tem prioridade.
    effective_path = file_path2 if file_path2 is not None else file_path
    if file_path2 is not None:
        print("[DEBUG] gerar_declaracao_escolar: usando file_path2 =", effective_path)
    else:
        print("[DEBUG] gerar_declaracao_escolar: usando file_path  =", effective_path)

    # HTML da tabela de notas (usado apenas na transferência do Fundamental)
    notas_tabela_html = ""

    # ------------------------------------------------------
    # 1) CARREGAMENTO DOS DADOS DO ALUNO (FUNDAMENTAL x EJA)
    # ------------------------------------------------------
    if session.get("declaracao_tipo") != "EJA":
        # ---------- FUNDAMENTAL ----------
        planilha = pd.read_excel(effective_path, sheet_name="LISTA CORRIDA")
        planilha.columns = [c.strip().upper() for c in planilha.columns]

        def format_rm(x):
            try:
                return str(int(float(x)))
            except Exception:
                return str(x)

        planilha["RM_str"] = planilha["RM"].apply(format_rm)

        try:
            rm_num = str(int(float(rm)))
        except Exception:
            rm_num = str(rm)

        aluno = planilha[planilha["RM_str"] == rm_num]
        if aluno.empty:
            return None

        row = aluno.iloc[0]

        semestre_texto = ""  # Fundamental não usa semestre
        nome = row["NOME"]
        serie = row["SÉRIE"]

        if isinstance(serie, str):
            # transforma "5ºA" em "5º ano A"
            serie = re.sub(r"(\d+º)([A-Za-z])", r"\1 ano \2", serie)

        data_nasc = row["DATA NASC."]
        ra = row["RA"]
        horario = row.get("HORÁRIO", "Desconhecido")

        if pd.isna(horario) or not str(horario).strip():
            horario = "Desconhecido"
        else:
            horario = str(horario).strip()

        ra_label = "RA"

        if pd.notna(data_nasc):
            try:
                data_nasc = pd.to_datetime(data_nasc, errors="coerce")
                if pd.notna(data_nasc):
                    data_nasc = data_nasc.strftime("%d/%m/%Y")
                else:
                    data_nasc = "Desconhecida"
            except Exception:
                data_nasc = "Desconhecida"
        else:
            data_nasc = "Desconhecida"

        # --------------------------------------------------
        # NOVO: busca as notas na aba NOTAS (apenas Transferência/Fundamental)
        # --------------------------------------------------
        if tipo == "Transferencia":
            try:
                notas_df = pd.read_excel(effective_path, sheet_name="NOTAS")
                notas_df.columns = [str(c).strip().upper() for c in notas_df.columns]

                # Descobre a coluna de RM (idealmente é "RM")
                rm_col = None
                for c in notas_df.columns:
                    if str(c).strip().upper() == "RM":
                        rm_col = c
                        break
                if rm_col is None and len(notas_df.columns) >= 3:
                    # fallback: coluna C (índice 2)
                    rm_col = notas_df.columns[2]

                if rm_col is not None:
                    notas_df["RM_str"] = notas_df[rm_col].apply(format_rm)
                    notas_aluno = notas_df[notas_df["RM_str"] == rm_num]

                    if not notas_aluno.empty:
                        notas_row = notas_aluno.iloc[0]

                        # devolve (texto, cor) – <5 vermelho, >=5 azul, SEMPRE com 2 casas decimais
                        def _fmt_nota(v):
                            if pd.isna(v):
                                return "—", None
                            s = str(v).strip()
                            if s == "":
                                return "—", None
                            try:
                                f = float(s.replace(",", "."))
                            except Exception:
                                # não conseguiu converter em número → sem cor especial
                                return s, None

                            # define cor
                            cor = "red" if f < 5 else "blue"

                            # formata SEMPRE com 2 casas decimais
                            texto = f"{f:.2f}".replace(".", ",")

                            return texto, cor

                        materias = [
                            ("Língua Portuguesa", "LP_1T", "LP_2T", "LP_3T"),
                            ("História", "HIST_1T", "HIST_2T", "HIST_3T"),
                            ("Geografia", "GEO_1T", "GEO_2T", "GEO_3T"),
                            ("Matemática", "MAT_1T", "MAT_2T", "MAT_3T"),
                            ("Ciências", "CIEN_1T", "CIEN_2T", "CIEN_3T"),
                            ("Educação Física", "EDFIS_1T", "EDFIS_2T", "EDFIS_3T"),
                            ("Arte", "ARTE_1T", "ARTE_2T", "ARTE_3T"),
                        ]

                        linhas_notas = ""

                        for nome_disc, c1, c2, c3 in materias:
                            n1_txt, n1_cor = _fmt_nota(notas_row.get(c1))
                            n2_txt, n2_cor = _fmt_nota(notas_row.get(c2))
                            n3_txt, n3_cor = _fmt_nota(notas_row.get(c3))

                            style_n1 = "border:1px solid #444;padding:3px 6px;text-align:center;"
                            style_n2 = "border:1px solid #444;padding:3px 6px;text-align:center;"
                            style_n3 = "border:1px solid #444;padding:3px 6px;text-align:center;"

                            if n1_cor:
                                style_n1 += f"color:{n1_cor};"
                            if n2_cor:
                                style_n2 += f"color:{n2_cor};"
                            if n3_cor:
                                style_n3 += f"color:{n3_cor};"

                            linhas_notas += (
                                "<tr>"
                                f"<td style='border:1px solid #444;padding:3px 6px;text-align:left;'>{nome_disc}</td>"
                                f"<td style='{style_n1}'>{n1_txt}</td>"
                                f"<td style='{style_n2}'>{n2_txt}</td>"
                                f"<td style='{style_n3}'>{n3_txt}</td>"
                                "</tr>"
                            )

                        if linhas_notas:
                            notas_tabela_html = (
                                "<br>"
                                "<span style='font-size:12px;'>"
                                "<strong>As notas do aluno, por componente curricular, são:</strong>"
                                "</span>"
                                "<br>"
                                "<table style='width:85%;max-width:700px;margin:4px auto 0 auto;"
                                "border-collapse:collapse;font-size:11px;'>"
                                "<thead>"
                                "<tr>"
                                "<th style='border:1px solid #444;padding:3px 6px;text-align:left;'>Componente curricular</th>"
                                "<th style='border:1px solid #444;padding:3px 6px;text-align:center;'>1º trim.</th>"
                                "<th style='border:1px solid #444;padding:3px 6px;text-align:center;'>2º trim.</th>"
                                "<th style='border:1px solid #444;padding:3px 6px;text-align:center;'>3º trim.</th>"
                                "</tr>"
                                "</thead>"
                                "<tbody>"
                                f"{linhas_notas}"
                                "</tbody>"
                                "</table>"
                            )
            except Exception as e:
                print(f"[ERRO] Falha ao carregar notas do aluno (Transferência Fundamental): {e}")
                notas_tabela_html = ""
    else:
        # ---------- EJA ----------
        df = pd.read_excel(effective_path, sheet_name=0, header=None, skiprows=1)
        df.columns = [str(c).strip().upper() for c in df.columns]

        # RM (coluna 2)
        df["RM_str"] = df.iloc[:, 2].apply(
            lambda x: str(int(x)) if pd.notna(x) and float(x) != 0 else ""
        )
        # Nome (coluna 3)
        df["NOME"] = df.iloc[:, 3]
        # Nascimento (coluna 6)
        df["NASC."] = df.iloc[:, 6]

        def get_ra(row_local):
            try:
                val = row_local.iloc[7]
                if pd.isna(val) or float(val) == 0:
                    return row_local.iloc[8]
                else:
                    return val
            except Exception:
                return row_local.iloc[7]

        df["RA"] = df.apply(get_ra, axis=1)
        # Série (coluna 0)
        df["SÉRIE"] = df.iloc[:, 0]

        try:
            rm_num = str(int(float(rm)))
        except Exception:
            rm_num = str(rm)

        aluno = df[df["RM_str"] == rm_num]
        if aluno.empty:
            return None

        row = aluno.iloc[0]

        # Semestre (quando existir na planilha)
        if len(row) > 29:
            semestre = row.iloc[29]
            semestre_texto = str(semestre).strip() if pd.notna(semestre) else ""
        else:
            semestre_texto = ""

        nome = row["NOME"]
        serie = row["SÉRIE"]
        if isinstance(serie, str):
            serie = re.sub(r"(\d+º)([A-Za-z])", r"\1 ano \2", serie)

        data_nasc = row["NASC."]
        ra = row["RA"]
        original_ra = row.iloc[7]

        # Se RA for vazio / 0, trata como RG
        if pd.isna(original_ra) or (
            isinstance(original_ra, (int, float)) and float(original_ra) == 0
        ):
            ra_label = "RG"
        else:
            ra_label = "RA"

        if pd.notna(data_nasc):
            try:
                data_nasc = pd.to_datetime(data_nasc, errors="coerce")
                data_nasc = (
                    data_nasc.strftime("%d/%m/%Y") if pd.notna(data_nasc) else "Desconhecida"
                )
            except Exception:
                data_nasc = "Desconhecida"
        else:
            data_nasc = "Desconhecida"

    # ------------------------------------------------------
    # 2) DATA POR EXTENSO
    # ------------------------------------------------------
    now = datetime.now()
    meses = {
        1: "janeiro",
        2: "fevereiro",
        3: "março",
        4: "abril",
        5: "maio",
        6: "junho",
        7: "julho",
        8: "agosto",
        9: "setembro",
        10: "outubro",
        11: "novembro",
        12: "dezembro",
    }
    mes_nome = meses[now.month].capitalize()
    data_extenso_str = f"Praia Grande, {now.day:02d} de {mes_nome} de {now.year}"

    additional_css = """
.print-button {
  background-color: #283E51;
  color: #fff;
  border: none;
  padding: 10px 20px;
  border-radius: 5px;
  cursor: pointer;
  margin-top: 20px;
}
.print-button:hover {
  background-color: #1d2d3a;
}
"""

    # ------------------------------------------------------
    # 3) MONTAGEM DO TEXTO DA DECLARAÇÃO
    # ------------------------------------------------------
    declaracao_text = ""
    tem_observacoes = False  # controla se há bloco de observações

    if tipo == "Escolaridade":
        titulo = "Declaração de Escolaridade"
        if session.get("declaracao_tipo") == "EJA":
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) "
                f"<strong><u>{nome}</u></strong>, portador(a) do {ra_label} "
                f"<strong><u>{ra}</u></strong>, nascido(a) em "
                f"<strong><u>{data_nasc}</u></strong>, "
                f"encontra-se regularmente matriculado(a) no segmento de "
                f"<strong><u>Educação de Jovens e Adultos (EJA)</u></strong> da "
                f"E.M José Padin Mouta, cursando atualmente o(a) "
                f"<strong><u>{serie}</u></strong>."
            )
        else:
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) "
                f"<strong><u>{nome}</u></strong>, portador(a) do RA "
                f"<strong><u>{ra}</u></strong>, nascido(a) em "
                f"<strong><u>{data_nasc}</u></strong>, "
                f"encontra-se regularmente matriculado(a) na "
                f"E.M José Padin Mouta, cursando atualmente o(a) "
                f"<strong><u>{serie}</u></strong> no horário de aula: "
                f"<strong><u>{horario}</u></strong>."
            )

    elif tipo == "Transferencia":
        titulo = "Declaração de Transferência"
        if session.get("declaracao_tipo") == "EJA":
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) "
                f"<strong><u>{nome}</u></strong>, portador(a) do {ra_label} "
                f"<strong><u>{ra}</u></strong>, nascido(a) em "
                f"<strong><u>{data_nasc}</u></strong>, matriculado(a) no segmento de "
                f"<strong><u>Educação de Jovens e Adultos (EJA)</u></strong> da "
                f"E.M José Padin Mouta, solicitou transferência desta unidade escolar "
                f"na data de hoje, estando apto(a) a cursar o(a) "
                f"<strong><u>{serie}</u></strong>."
            )
        else:
            serie_mod = re.sub(r"^(\d+º).*", r"\1 ano", str(serie))
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) responsável do(a) "
                f"aluno(a) <strong><u>{nome}</u></strong>, portador(a) do RA "
                f"<strong><u>{ra}</u></strong>, nascido(a) em "
                f"<strong><u>{data_nasc}</u></strong>, compareceu a nossa "
                f"unidade escolar e solicitou transferência na data de hoje, "
                f"o aluno está apto(a) a cursar o(a) "
                f"<strong><u>{serie_mod}</u></strong>."
            )
            # Anexa a tabela de notas (apenas Fundamental)
            if notas_tabela_html:
                declaracao_text += notas_tabela_html

    elif tipo == "Conclusão":
        titulo = "Declaração de Conclusão"

        if session.get("declaracao_tipo") == "EJA":
            # Próxima série/etapa
            mapping = {
                "1ª SÉRIE E.F": "2ª SÉRIE E.F",
                "2ª SÉRIE E.F": "3ª SÉRIE E.F",
                "3ª SÉRIE E.F": "4ª SÉRIE E.F",
                "4ª SÉRIE E.F": "5ª SÉRIE E.F",
                "5ª SÉRIE E.F": "6ª SÉRIE E.F",
                "6ª SÉRIE E.F": "7ª SÉRIE E.F",
                "7ª SÉRIE E.F": "8ª SÉRIE E.F",
                "8ª SÉRIE E.F": "1ª SÉRIE E.M",
                "1ª SÉRIE E.M": "2ª SÉRIE E.M",
                "2ª SÉRIE E.M": "3ª SÉRIE E.M",
                "3ª SÉRIE E.M": "ENSINO SUPERIOR",
            }
            series_text = mapping.get(str(serie).upper(), "a série subsequente")

            # Se tiver semestre na planilha, só acrescenta como complemento
            semestre_parte = (
                f", no <strong><u>{semestre_texto}</u></strong>"
                if semestre_texto
                else ""
            )

            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) "
                f"<strong><u>{nome}</u></strong>, portador(a) do {ra_label} "
                f"<strong><u>{ra}</u></strong>, nascido(a) em "
                f"<strong><u>{data_nasc}</u></strong>, concluiu com êxito o(a) "
                f"<strong><u>{serie}</u></strong>{semestre_parte} no segmento de "
                f"<strong><u>Educação de Jovens e Adultos (EJA)</u></strong> da "
                f"E.M José Padin Mouta, estando apto(a) a ingressar no(na) "
                f"<strong><u>{series_text}</u></strong>."
            )
        else:
            match = re.search(r"(\d+)º\s*ano", str(serie))
            next_year = int(match.group(1)) + 1 if match else None
            series_text = f"{next_year}º ano" if next_year else "a série subsequente"

            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) "
                f"<strong><u>{nome}</u></strong>, portador(a) do RA "
                f"<strong><u>{ra}</u></strong>, nascido(a) em "
                f"<strong><u>{data_nasc}</u></strong>, concluiu com êxito o(a) "
                f"<strong><u>{serie}</u></strong>, estando apto(a) a ingressar no(na) "
                f"<strong><u>{series_text}</u></strong>."
            )

    elif tipo in ("Frequencia", "Frequência"):
        titulo = "Declaração de Frequência"

        # Se não vier dados de frequência, não há o que declarar
        if not dados_frequencia or not dados_frequencia.get("meses"):
            return None

        # Frase inicial (EJA x Fundamental)
        if session.get("declaracao_tipo") == "EJA":
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) "
                f"<strong><u>{nome}</u></strong>, portador(a) do {ra_label} "
                f"<strong><u>{ra}</u></strong>, nascido(a) em "
                f"<strong><u>{data_nasc}</u></strong>, regularmente matriculado(a) "
                f"no segmento de <strong><u>Educação de Jovens e Adultos (EJA)</u></strong> "
                f"da E.M José Padin Mouta, teve sua frequência apurada nos meses abaixo "
                f"conforme quadro a seguir."
            )
        else:
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) "
                f"<strong><u>{nome}</u></strong>, portador(a) do RA "
                f"<strong><u>{ra}</u></strong>, nascido(a) em "
                f"<strong><u>{data_nasc}</u></strong>, regularmente matriculado(a) "
                f"no(a) <strong><u>{serie}</u></strong> da E.M José Padin Mouta, "
                f"teve sua frequência apurada nos meses abaixo conforme quadro a seguir."
            )

        def _fmt_num(n):
            try:
                f = float(n)
                if f.is_integer():
                    return str(int(f))
                return f"{f:.1f}".replace(".", ",")
            except Exception:
                return str(n)

        # mapa padrão de meses (1–12)
        nomes_meses_padrao = {
            1: "Janeiro",
            2: "Fevereiro",
            3: "Março",
            4: "Abril",
            5: "Maio",
            6: "Junho",
            7: "Julho",
            8: "Agosto",
            9: "Setembro",
            10: "Outubro",
            11: "Novembro",
            12: "Dezembro",
        }

        linhas_tabela = ""

        for idx, item in enumerate(dados_frequencia.get("meses", []), start=1):
            # ---------- MÊS ----------
            raw_mes = item.get("nome_mes")
            if raw_mes in (None, ""):
                raw_mes = item.get("mes")
            if raw_mes in (None, ""):
                raw_mes = item.get("mes_nome")
            if raw_mes in (None, ""):
                raw_mes = item.get("descricao_mes")
            if raw_mes in (None, ""):
                raw_mes = item.get("descricao")

            nome_mes = ""

            if raw_mes not in (None, ""):
                if isinstance(raw_mes, (int, float)):
                    idx_int = int(raw_mes)
                    nome_mes = nomes_meses_padrao.get(idx_int, str(idx_int))
                else:
                    s = str(raw_mes).strip()
                    if s.isdigit():
                        idx_int = int(s)
                        nome_mes = nomes_meses_padrao.get(idx_int, s)
                    else:
                        nome_mes = s

            # fallback final: usa o índice do loop (1..12)
            if not nome_mes:
                nome_mes = nomes_meses_padrao.get(idx, f"Mês {idx}")

            # ---------- DIAS / FALTAS / FREQUÊNCIA ----------
            dias_val = item.get("dias_letivos_calculados")
            if dias_val is None:
                dias_val = item.get("dias_letivos")
            if dias_val is None:
                dias_val = item.get("dias")

            faltas_val = item.get("faltas_calculadas")
            if faltas_val is None:
                faltas_val = item.get("faltas")

            freq_val = item.get("frequencia")
            if freq_val is None:
                freq_val = item.get("freq")

            # Se não vier flag 'preenchido', deduz pelo conteúdo
            preenchido_raw = item.get("preenchido")
            if preenchido_raw is None:
                preenchido = any(
                    v not in (None, "", 0, 0.0)
                    for v in (dias_val, faltas_val, freq_val)
                )
            else:
                preenchido = bool(preenchido_raw)

            if preenchido:
                dias_txt = _fmt_num(dias_val) if dias_val not in (None, "") else "0"
                faltas_txt = _fmt_num(faltas_val) if faltas_val not in (None, "") else "0"
                if freq_val in (None, ""):
                    freq_txt = "—"
                else:
                    try:
                        freq_txt = f"{float(freq_val):.1f}%".replace(".", ",")
                    except Exception:
                        freq_txt = str(freq_val)
            else:
                dias_txt = "—"
                faltas_txt = "—"
                freq_txt = "—"

            linhas_tabela += (
                "<tr>"
                f"<td style='border:1px solid #444;padding:4px 6px;text-align:center;'>{nome_mes}</td>"
                f"<td style='border:1px solid #444;padding:4px 6px;text-align:center;'>{dias_txt}</td>"
                f"<td style='border:1px solid #444;padding:4px 6px;text-align:center;'>{faltas_txt}</td>"
                f"<td style='border:1px solid #444;padding:4px 6px;text-align:center;'>{freq_txt}</td>"
                "</tr>"
            )

        declaracao_text += (
            "<br><br>"
            "<table style='width:75%;max-width:600px;margin:0 auto;"
            "border-collapse:collapse;font-size:12px;margin-top:4px;'>"
            "<thead>"
            "<tr>"
            "<th style='border:1px solid #444;padding:4px 6px;text-align:center;'>Mês</th>"
            "<th style='border:1px solid #444;padding:4px 6px;text-align:center;'>Dias letivos</th>"
            "<th style='border:1px solid #444;padding:4px 6px;text-align:center;'>Faltas</th>"
            "<th style='border:1px solid #444;padding:4px 6px;text-align:center;'>Frequência</th>"
            "</tr>"
            "</thead>"
            "<tbody>"
            f"{linhas_tabela}"
            "</tbody>"
            "</table>"
            "<br>"
            "<span style='font-size:12px;color:#555;'>"
            "</span>"
        )

    else:
        # Tipo desconhecido
        return None

    # ------------------------------------------------------
    # 4) OBSERVAÇÕES / HISTÓRICO / BOLSA FAMÍLIA
    # ------------------------------------------------------
    valor_bolsa = str(row.get("BOLSA FAMILIA", "")).strip().upper()

    if deve_historico or (valor_bolsa == "SIM" and tipo != "Escolaridade"):
        tem_observacoes = True
        declaracao_text += "<br><br><strong>Observações:</strong><br>"
        declaracao_text += (
            '<label class="checkbox-label" '
            "style='display:block;text-align:justify;font-size:14px;'>"
        )

        # Histórico escolar pendente
        if deve_historico:
            declaracao_text += '<span class="warning-icon">&#9888;</span> '
            declaracao_text += (
                "O aluno deve o histórico escolar da unidade anterior:<br><br>"
            )

            if unidade_anterior:
                unidade_anterior = " ".join(str(unidade_anterior).strip().split())
                esc_df = None
                if escolas_df is not None:
                    try:
                        esc_df = escolas_df[
                            escolas_df.iloc[:, 3].str.upper()
                            == unidade_anterior.upper()
                        ]
                    except Exception:
                        esc_df = None

                if esc_df is not None and not esc_df.empty:
                    unidade_nome = esc_df.iloc[0, 3]
                    inep = esc_df.iloc[0, 4]
                    municipio = esc_df.iloc[0, 2]
                    uf = esc_df.iloc[0, 1]
                    declaracao_text += (
                        "<div style='font-size:14px;'>"
                        f"<strong>Unidade:</strong> {unidade_nome}<br>"
                        f"<strong>INEP:</strong> {inep}<br>"
                        f"<strong>Cidade:</strong> {municipio}<br>"
                        f"<strong>Estado:</strong> {uf}<br><br>"
                        "</div>"
                    )
                else:
                    declaracao_text += (
                        f"<strong>Unidade:</strong> {unidade_anterior}<br><br>"
                    )

            declaracao_text += (
                "Após sua entrega, o documento será confeccionado em "
                "até 30 dias úteis.<br><br>"
            )

        # Bolsa Família (quando não for só escolaridade)
        if valor_bolsa == "SIM" and tipo != "Escolaridade":
            declaracao_text += (
                '<img src="/static/logos/bolsa_familia.jpg" '
                'alt="Bolsa Família" '
                'style="width:28px;vertical-align:middle;margin-right:5px;">'
                "O aluno é beneficiário do Programa Bolsa Família."
            )

        declaracao_text += "</label>"

    # ------------------------------------------------------
    # 5) HTML FINAL (DECLARAÇÃO ÚNICA)
    # ------------------------------------------------------
    classes_body = []
    if tipo in ("Frequencia", "Frequência"):
        classes_body.append("tipo-frequencia")
    if tipo == "Transferencia" and tem_observacoes:
        classes_body.append("transferencia-com-observacoes")

    body_class_attr = f' class="{" ".join(classes_body)}"' if classes_body else ""

    base_template = f"""<!doctype html>
<html lang="pt-br">
<head>
  <meta charset="utf-8">
  <title>{titulo} - E.M José Padin Mouta</title>
  <style>
    @page {{
      margin: 0;
    }}
    html, body {{
      margin: 0;
      padding: 0.5cm;
      font-family: 'Montserrat', sans-serif;
      font-size: 16px;
      line-height: 1.5;
      color: #333;
    }}
    .header {{
      text-align: center;
      border-bottom: 2px solid #283E51;
      padding-bottom: 5px;
      margin-bottom: 10px;
    }}
    .header h1 {{
      margin: 0;
      font-size: 24px;
      text-transform: uppercase;
      color: #283E51;
    }}
    .header p {{
      margin: 3px 0;
      font-size: 16px;
    }}
    .date {{
      text-align: right;
      font-size: 16px;
      margin-bottom: 10px;
    }}
    .content {{
      text-align: justify;
      margin-bottom: 10px;
      padding: 0 2cm;
      box-sizing: border-box;
      hyphens: auto;
      word-wrap: break-word;
      overflow-wrap: break-word;
    }}
    .content p {{
      white-space: normal !important;
      word-break: break-word !important;
      overflow-wrap: break-word !important;
      hyphens: auto !important;
    }}
    .signature {{
      text-align: center;
      margin: 0;
      padding: 0;
    }}
    .signature .line {{
      height: 1px;
      background-color: #333;
      width: 60%;
      margin: 0 auto 5px auto;
    }}
    .footer {{
      text-align: center;
      border-top: 2px solid #283E51;
      padding-top: 5px;
      margin: 0;
      font-size: 14px;
      color: #555;
    }}
    .warning-icon {{
      display: inline-block;
      width: 18px;
      height: 18px;
      color: red;
      font-weight: bold;
      font-size: 18px;
      line-height: 18px;
      vertical-align: middle;
      user-select: none;
    }}

    /* Ajustes específicos apenas para a declaração de Frequência (tela) */
    body.tipo-frequencia {{
      font-size: 14px;
    }}
    body.tipo-frequencia .content {{
      padding: 0 1.5cm;
      margin-bottom: 8px;
    }}
    body.tipo-frequencia table {{
      font-size: 14px;
    }}
    body.tipo-frequencia .footer {{
      font-size: 14px;
    }}
    body.tipo-frequencia .signature p {{
      font-size: 14px;
    }}

    @media print {{
      .no-print {{ display: none !important; }}
      body {{
        margin: 0;
        padding: 0.5cm 0.5cm;
        font-size: 14px;
        font-family: 'Montserrat', sans-serif;
        color: #000;
      }}
      .declaration-bottom {{
        position: fixed;
        bottom: 0;
        left: 0;
        width: 100%;
      }}
      .date {{
        margin: 1cm 0 1cm 0;
        text-align: right;
        hyphens: none !important;
      }}
      .content {{
        margin: 0;
        padding: 0;
        text-align: justify !important;
        hyphens: none !important;
        white-space: normal !important;
        word-wrap: break-word !important;
        overflow-wrap: break-word !important;
      }}
      .content p {{
        margin: 0 0 1em 0;
        text-align: justify !important;
        hyphens: none !important;
      }}
      body, .content, .content p {{
        width: auto !important;
        max-width: none !important;
      }}

      /* Override APENAS na declaração de frequência */
      body.tipo-frequencia {{
        padding: 0.3cm 0.3cm;
        font-size: 14px;
      }}
      body.tipo-frequencia .content {{
        margin: 0 0 0.3cm 0;
        padding: 0;
      }}
      body.tipo-frequencia .declaration-bottom {{
        position: static;
        margin-top: 1.0cm;
      }}
      body.tipo-frequencia .footer,
      body.tipo-frequencia .signature p {{
        font-size: 14px;
      }}

      /* Transferência com observações */
      body.transferencia-com-observacoes {{
        padding: 1.2cm 1.2cm;
        font-size: 14px;
      }}
      body.transferencia-com-observacoes .content {{
        margin: 0 0 0.6cm 0;
        padding: 0;
      }}
      body.transferencia-com-observacoes .declaration-bottom {{
        position: static;
        margin-top: 1.2cm;
      }}
      body.transferencia-com-observacoes .footer,
      body.transferencia-com-observacoes .signature p {{
        font-size: 14px;
      }}
    }}

    .content, .content p, .date {{
      hyphens: none !important;
      word-break: normal !important;
      overflow-wrap: normal !important;
    }}

    {additional_css}

    .checkbox-label {{
      display: flex;
      align-items: center;
      gap: 8px;
      text-align: left !important;
      font-size: 14px;
      margin-top: 8px;
      margin-bottom: 8px;
      flex-wrap: wrap;
    }}

    header {{
      background: linear-gradient(90deg, #283E51, #4B79A1);
      color: #fff;
      padding: 20px;
      text-align: center;
      border-bottom: 3px solid #1d2d3a;
      border-radius: 0 0 15px 15px;
      box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }}
  </style>
</head>
<body{body_class_attr}>
  <div class="declaration-container">
    <div class="header">
      <div style="display:flex;justify-content:space-between;align-items:center;">
        <img src="/static/logos/escola.png" alt="Escola Logo" style="height:80px;">
        <div>
          <h1>Secretaria de Educação</h1>
          <p>E.M José Padin Mouta</p>
          <p>Município da Estância Balneária de Praia Grande</p>
          <p>Estado de São Paulo</p>
        </div>
        <img src="/static/logos/municipio.png" alt="Município Logo" style="height:80px;">
      </div>
    </div>
    <div class="date">
      <p>{data_extenso_str}</p>
    </div>
    <div class="content">
      <h2 style="text-align:center;text-transform:uppercase;color:#283E51;">
        {titulo}
      </h2>
      <p>{declaracao_text}</p>
    </div>
    <div class="declaration-bottom">
      <div class="signature">
        <div class="line"></div>
        <p>Luciana Rocha Augustinho</p>
        <p>Diretora da Unidade Escolar</p>
      </div>
      <div class="footer">
        <p>Rua: Bororós, nº 150, Vila Tupi, Praia Grande - SP - Telefone: 3496-5321 | E-mail: em.padin@praiagrande.sp.gov.br</p>
      </div>
    </div>
  </div>
  <div class="no-print" style="text-align:center;margin-top:20px;">
    <button onclick="window.print()" class="print-button">Imprimir Declaração</button>
  </div>
</body>
</html>
"""
    return base_template


# ==========================================================
#  NOVA FUNÇÃO – LOTE DE ESCOLARIDADE 5º ANO (FUNDAMENTAL)
# ==========================================================

def gerar_lote_escolaridade_5ano(file_path, file_path2=None):
    """
    Gera os dados para DECLARAÇÕES DE ESCOLARIDADE de todos os alunos de 5º ano
    (Fundamental) em lote.
    """
    effective_path = file_path2 if file_path2 is not None else file_path
    if not effective_path:
        raise ValueError(
            "Caminho do arquivo Excel não informado para o lote de escolaridade 5º ano."
        )

    planilha = pd.read_excel(effective_path, sheet_name="LISTA CORRIDA")
    planilha.columns = [c.strip().upper() for c in planilha.columns]

    def format_rm(x):
        try:
            return str(int(float(x)))
        except Exception:
            return str(x)

    planilha["RM_str"] = planilha["RM"].apply(format_rm)

    registros = []

    for _, row in planilha.iterrows():
        rm_str = str(row.get("RM_str", "")).strip()
        if rm_str in ("", "0"):
            continue

        serie_raw = str(row.get("SÉRIE", "")).strip()
        if not serie_raw:
            continue

        if "5º" not in serie_raw and "5°" not in serie_raw:
            continue

        nome = str(row.get("NOME", "")).strip()
        ra = str(row.get("RA", "")).strip()

        data_nasc_val = row.get("DATA NASC.")
        if pd.notna(data_nasc_val):
            try:
                data_nasc_dt = pd.to_datetime(data_nasc_val, errors="coerce")
                data_nasc = (
                    data_nasc_dt.strftime("%d/%m/%Y")
                    if pd.notna(data_nasc_dt)
                    else "Desconhecida"
                )
            except Exception:
                data_nasc = "Desconhecida"
        else:
            data_nasc = "Desconhecida"

        horario = str(row.get("HORÁRIO", "")).strip()
        if not horario:
            horario = "Desconhecido"

        serie_fmt = serie_raw
        try:
            serie_fmt = re.sub(r"(\d+º)\s*([A-Za-z])", r"\1 ano \2", serie_fmt)
        except Exception:
            pass

        texto = (
            f"Declaro, para os devidos fins, que o(a) aluno(a) "
            f"<strong><u>{nome}</u></strong>, portador(a) do RA "
            f"<strong><u>{ra}</u></strong>, nascido(a) em "
            f"<strong><u>{data_nasc}</u></strong>, "
            f"encontra-se regularmente matriculado(a) na "
            f"E.M José Padin Mouta, cursando atualmente o(a) "
            f"<strong><u>{serie_fmt}</u></strong> no horário de aula: "
            f"<strong><u>{horario}</u></strong>."
        )

        registros.append(
            {
                "nome": nome,
                "ra": ra,
                "data_nasc": data_nasc,
                "serie_fmt": serie_fmt,
                "horario": horario,
                "texto": texto,
            }
        )

    now = datetime.now()
    meses = {
        1: "janeiro",
        2: "fevereiro",
        3: "março",
        4: "abril",
        5: "maio",
        6: "junho",
        7: "julho",
        8: "agosto",
        9: "setembro",
        10: "outubro",
        11: "novembro",
        12: "dezembro",
    }
    mes_nome = meses[now.month].capitalize()
    data_extenso_str = f"Praia Grande, {now.day:02d} de {mes_nome} de {now.year}"
    titulo = "Declaração de Escolaridade"

    return registros, data_extenso_str, titulo


# ==========================================================
#  DECLARAÇÃO PERSONALIZADA (Fundamental / EJA)
# ==========================================================

def gerar_declaracao_personalizada(dados):
    """
    Gera o HTML de declarações personalizadas (Conclusão, Matrícula cancelada
    ou Não Comparecimento - NCOM).
    """

    def _get_str(key, default=""):
        return (dados.get(key) or default).strip()

    def _normalizar_semestre(*keys):
        for k in keys:
            v = dados.get(k)
            if v is None:
                continue
            s = str(v).strip()
            if s:
                return s
        return ""

    nome = _get_str("nome_aluno")
    ra = _get_str("ra")
    data_nasc_raw = _get_str("data_nascimento")

    seg_raw = dados.get("segmento") or dados.get("segmento_personalizado") or "Fundamental"
    seg_norm = str(seg_raw).strip().lower()
    if seg_norm in ("fundamental", "fund", "ef", "ensino fundamental"):
        segmento = "Fundamental"
    else:
        segmento = "EJA"

    data_nasc = "Desconhecida"
    if data_nasc_raw:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                dt = datetime.strptime(data_nasc_raw, fmt)
                data_nasc = dt.strftime("%d/%m/%Y")
                break
            except Exception:
                continue

    if segmento == "Fundamental":
        segmento_label = "Ensino Fundamental"
        prep_segmento = "do"
    else:
        segmento_label = "Educação de Jovens e Adultos (EJA)"
        prep_segmento = "da"

    tipo_decl_raw = dados.get("tipo_declaracao") or dados.get("tipo_declaracao_personalizada")
    tipo_decl = (tipo_decl_raw or "").strip().lower()

    declaracao_text = ""
    titulo = ""

    if tipo_decl in ("conclusao", "conclusão"):
        titulo = "Declaração de Conclusão"
        ano_serie = _get_str("ano_serie_concluida")
        ano_conclusao = _get_str("ano_conclusao")

        deve_hist_val_raw = dados.get("deve_historico_unidade")
        deve_hist_str = str(deve_hist_val_raw or "").strip().lower()
        deve_hist_unidade = deve_hist_str in ("sim", "1", "true", "on")

        if segmento == "Fundamental":
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) "
                f"<strong><u>{nome}</u></strong>, portador(a) do RA "
                f"<strong><u>{ra}</u></strong>, nascido(a) em "
                f"<strong><u>{data_nasc}</u></strong>, concluiu o(a) "
                f"<strong><u>{ano_serie}</u></strong> {prep_segmento} "
                f"<strong><u>{segmento_label}</u></strong>, no ano letivo de "
                f"<strong><u>{ano_conclusao}</u></strong>, nesta unidade escolar."
            )
        else:
            semestre_conclusao = _normalizar_semestre(
                "semestre_conclusao",
                "semestre_conclusao_opcao",
                "semestre_matricula",
                "semestre_matricula_opcao",
            )

            if semestre_conclusao:
                periodo_conclusao = (
                    f"no <strong><u>{semestre_conclusao}</u></strong> do ano de "
                    f"<strong><u>{ano_conclusao}</u></strong>"
                )
            else:
                periodo_conclusao = (
                    f"no ano letivo de <strong><u>{ano_conclusao}</u></strong>"
                )

            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) "
                f"<strong><u>{nome}</u></strong>, portador(a) do RA "
                f"<strong><u>{ra}</u></strong>, nascido(a) em "
                f"<strong><u>{data_nasc}</u></strong>, concluiu o(a) "
                f"<strong><u>{ano_serie}</u></strong> no segmento de "
                f"<strong><u>Educação de Jovens e Adultos (EJA)</u></strong>, "
                f"{periodo_conclusao}, nesta unidade escolar."
            )

        if deve_hist_unidade:
            declaracao_text += (
                " Ressalta-se que consta, junto a esta unidade escolar, "
                "pendência de histórico escolar referente ao(à) aluno(a) citado(a)."
            )

    elif tipo_decl in ("matriculacancelada", "matricula cancelada", "matricula_cancelada"):
        titulo = "Declaração de Matrícula Cancelada"
        ano_serie = _get_str("ano_serie_matricula")
        ano_matricula = _get_str("ano_matricula")
        semestre_matricula = _normalizar_semestre(
            "semestre_matricula",
            "semestre_matricula_opcao",
        )

        if segmento == "EJA" and semestre_matricula:
            periodo_matricula = (
                f"no <strong><u>{semestre_matricula}</u></strong> do ano de "
                f"<strong><u>{ano_matricula}</u></strong>"
            )
        else:
            periodo_matricula = (
                f"no ano de <strong><u>{ano_matricula}</u></strong>"
            )

        if segmento == "EJA":
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) "
                f"<strong><u>{nome}</u></strong>, portador(a) do RA "
                f"<strong><u>{ra}</u></strong>, nascido(a) em "
                f"<strong><u>{data_nasc}</u></strong>, esteve matriculado(a) no(a) "
                f"<strong><u>{ano_serie}</u></strong> no segmento de "
                f"<strong><u>Educação de Jovens e Adultos (EJA)</u></strong>, "
                f"{periodo_matricula}, nesta unidade escolar, tendo sua matrícula cancelada."
            )
        else:
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) "
                f"<strong><u>{nome}</u></strong>, portador(a) do RA "
                f"<strong><u>{ra}</u></strong>, nascido(a) em "
                f"<strong><u>{data_nasc}</u></strong>, esteve matriculado(a) no(a) "
                f"<strong><u>{ano_serie}</u></strong> {prep_segmento} "
                f"<strong><u>{segmento_label}</u></strong>, {periodo_matricula}, "
                "nesta unidade escolar, tendo sua matrícula cancelada."
            )

    elif tipo_decl == "ncom":
        titulo = "Declaração de Não Comparecimento (NCOM)"
        ano_serie = _get_str("ano_serie_vaga")
        ano_ref = _get_str("ano_referencia_ncom")
        semestre_ref = _normalizar_semestre(
            "semestre_referencia_ncom",
            "semestre_referencia",
        )

        if segmento == "EJA" and semestre_ref:
            periodo_ref = (
                f"para o <strong><u>{semestre_ref}</u></strong> do ano de "
                f"<strong><u>{ano_ref}</u></strong>"
            )
        else:
            periodo_ref = (
                f"para o ano de <strong><u>{ano_ref}</u></strong>"
            )

        if segmento == "EJA":
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) "
                f"<strong><u>{nome}</u></strong>, portador(a) do RA "
                f"<strong><u>{ra}</u></strong>, nascido(a) em "
                f"<strong><u>{data_nasc}</u></strong>, teve vaga destinada ao(à) "
                f"<strong><u>{ano_serie}</u></strong> no segmento de "
                f"<strong><u>Educação de Jovens e Adultos (EJA)</u></strong>, "
                f"{periodo_ref} nesta unidade escolar. Todavia, o(a) aluno(a) "
                "não compareceu à unidade escolar, sendo considerado(a) NCOM – "
                "Não Comparecimento, motivo pelo qual a vaga foi cancelada nesta "
                "unidade escolar."
            )
        else:
            declaracao_text = (
                f"Declaro, para os devidos fins, que o(a) aluno(a) "
                f"<strong><u>{nome}</u></strong>, portador(a) do RA "
                f"<strong><u>{ra}</u></strong>, nascido(a) em "
                f"<strong><u>{data_nasc}</u></strong>, teve vaga destinada ao(à) "
                f"<strong><u>{ano_serie}</u></strong> {prep_segmento} "
                f"<strong><u>{segmento_label}</u></strong>, {periodo_ref} "
                "nesta unidade escolar. Todavia, o(a) aluno(a) não compareceu à unidade "
                "escolar, sendo considerado(a) NCOM – Não Comparecimento, motivo pelo qual "
                "a vaga foi cancelada nesta unidade escolar."
            )

    else:
        return None

    now = datetime.now()
    meses = {
        1: "janeiro",
        2: "fevereiro",
        3: "março",
        4: "abril",
        5: "maio",
        6: "junho",
        7: "julho",
        8: "agosto",
        9: "setembro",
        10: "outubro",
        11: "novembro",
        12: "dezembro",
    }
    mes = meses[now.month].capitalize()
    data_extenso_str = f"Praia Grande, {now.day:02d} de {mes} de {now.year}"

    additional_css = """
.print-button {
  background-color: #283E51;
  color: #fff;
  border: none;
  padding: 10px 20px;
  border-radius: 5px;
  cursor: pointer;
  margin-top: 20px;
}
.print-button:hover {
  background-color: #1d2d3a;
}
"""

    base_template = f"""<!doctype html>
<html lang="pt-br">
<head>
  <meta charset="utf-8">
  <title>{titulo} - E.M José Padin Mouta</title>
  <style>
    @page {{
      margin: 0;
    }}
    html, body {{
      margin: 0;
      padding: 0.5cm;
      font-family: 'Montserrat', sans-serif;
      font-size: 16px;
      line-height: 1.5;
      color: #333;
    }}
    .header {{
      text-align: center;
      border-bottom: 2px solid #283E51;
      padding-bottom: 5px;
      margin-bottom: 10px;
    }}
    .header h1 {{
      margin: 0;
      font-size: 24px;
      text-transform: uppercase;
      color: #283E51;
    }}
    .header p {{
      margin: 3px 0;
      font-size: 16px;
    }}
    .date {{
      text-align: right;
      font-size: 16px;
      margin-bottom: 10px;
    }}
    .content {{
      text-align: justify;
      margin-bottom: 10px;
      padding: 0 2cm;
      box-sizing: border-box;
      hyphens: auto;
      word-wrap: break-word;
      overflow-wrap: break-word;
    }}
    .content p {{
      white-space: normal !important;
      word-break: break-word !important;
      overflow-wrap: break-word !important;
      hyphens: auto !important;
    }}
    .signature {{
      text-align: center;
      margin: 0;
      padding: 0;
    }}
    .signature .line {{
      height: 1px;
      background-color: #333;
      width: 60%;
      margin: 0 auto 5px auto;
    }}
    .footer {{
      text-align: center;
      border-top: 2px solid #283E51;
      padding-top: 5px;
      margin: 0;
      font-size: 14px;
      color: #555;
    }}
    .warning-icon {{
      display: inline-block;
      width: 18px;
      height: 18px;
      color: red;
      font-weight: bold;
      font-size: 18px;
      line-height: 18px;
      vertical-align: middle;
      user-select: none;
    }}

    @media print {{
      .no-print {{ display: none !important; }}
      body {{
        margin: 0;
        padding: 1.5cm 1.5cm;
        font-size: 14px;
        font-family: 'Montserrat', sans-serif;
        color: #000;
      }}
      .declaration-bottom {{
        position: fixed;
        bottom: 0;
        left: 0;
        width: 100%;
      }}
      .date {{
        margin: 1cm 0 1cm 0;
        text-align: right;
        hyphens: none !important;
      }}
      .content {{
        margin: 0;
        padding: 0;
        text-align: justify !important;
        hyphens: none !important;
        white-space: normal !important;
        word-wrap: break-word !important;
        overflow-wrap: break-word !important;
      }}
      .content p {{
        margin: 0 0 1em 0;
        text-align: justify !important;
        hyphens: none !important;
      }}
      body, .content, .content p {{
        width: auto !important;
         max-width: none !important;
      }}
    }}

    .content, .content p, .date {{
      hyphens: none !important;
      word-break: normal !important;
      overflow-wrap: normal !important;
    }}

    {additional_css}

    .checkbox-label {{
      display: flex;
      align-items: center;
      gap: 8px;
      text-align: left !important;
      font-size: 14px;
      margin-top: 8px;
      margin-bottom: 8px;
      flex-wrap: wrap;
    }}

    header {{
      background: linear-gradient(90deg, #283E51, #4B79A1);
      color: #fff;
      padding: 20px;
      text-align: center;
      border-bottom: 3px solid #1d2d3a;
      border-radius: 0 0 15px 15px;
      box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }}
  </style>
</head>
<body>
  <div class="declaration-container">
    <div class="header">
      <div style="display:flex;justify-content:space-between;align-items:center;">
        <img src="/static/logos/escola.png" alt="Escola Logo" style="height:80px;">
        <div>
          <h1>Secretaria de Educação</h1>
          <p>E.M José Padin Mouta</p>
          <p>Município da Estância Balneária de Praia Grande</p>
          <p>Estado de São Paulo</p>
        </div>
        <img src="/static/logos/municipio.png" alt="Município Logo" style="height:80px;">
      </div>
    </div>
    <div class="date">
      <p>{data_extenso_str}</p>
    </div>
    <div class="content">
      <h2 style="text-align:center;text-transform:uppercase;color:#283E51;">
        {titulo}
      </h2>
      <p>{declaracao_text}</p>
    </div>
    <div class="declaration-bottom">
      <div class="signature">
        <div class="line"></div>
        <p>Luciana Rocha Augustinho</p>
        <p>Diretora da Unidade Escolar</p>
      </div>
      <div class="footer">
        <p>Rua: Bororós, nº 150, Vila Tupi, Praia Grande - SP - Telefone: 3496-5321 | E-mail: em.padin@praiagrande.sp.gov.br</p>
      </div>
    </div>
  </div>
  <div class="no-print" style="text-align:center;margin-top:20px;">
    <button onclick="window.print()" class="print-button">Imprimir Declaração</button>
  </div>
</body>
</html>
"""
    return base_template


# ==========================================================
#  AUTENTICAÇÃO / DASHBOARD
# ==========================================================

@app.route("/login", methods=["GET", "POST"])
def login_route():
    error = None
    if request.method == "POST":
        token = request.form.get("token")
        if token == ACCESS_TOKEN:
            session["logged_in"] = True

            # ==========================================================
            # AJUSTE: EJA NÃO É MAIS OBRIGATÓRIA
            # - Só exige lista_fundamental para seguir para o dashboard
            # - lista_eja permanece suportada, mas opcional
            # ==========================================================
            if "lista_fundamental" not in session:
                return redirect(url_for("upload_listas"))

            return redirect(url_for("dashboard"))
        else:
            error = "Token inválido. Tente novamente."

    return render_template("login.html", error=error)


@app.route("/logout")
def logout_route():
    session.clear()
    return redirect(url_for("login_route"))


@app.route("/upload_listas", methods=["GET", "POST"])
@login_required
def upload_listas():
    if request.method == "POST":
        fundamental_file = request.files.get("lista_fundamental")
        eja_file = request.files.get("lista_eja")  # agora é opcional

        if not fundamental_file or fundamental_file.filename == "":
            flash("Selecione a Lista Piloto - REGULAR - 2025", "error")
            return redirect(url_for("upload_listas"))

        # Salva Fundamental (obrigatório)
        fundamental_filename = secure_filename(
            f"fundamental_{uuid.uuid4().hex}_" + fundamental_file.filename
        )
        fundamental_path = os.path.join(app.config["UPLOAD_FOLDER"], fundamental_filename)
        fundamental_file.save(fundamental_path)
        session["lista_fundamental"] = fundamental_path

        # Salva EJA (opcional)
        eja_salva = False
        if eja_file and eja_file.filename:
            eja_filename = secure_filename(f"eja_{uuid.uuid4().hex}_" + eja_file.filename)
            eja_path = os.path.join(app.config["UPLOAD_FOLDER"], eja_filename)
            eja_file.save(eja_path)
            session["lista_eja"] = eja_path
            eja_salva = True

        if eja_salva:
            flash("Listas carregadas com sucesso (Fundamental e EJA).", "success")
        else:
            flash(
                "Lista do Fundamental carregada com sucesso. A lista de EJA é opcional e não foi enviada.",
                "success",
            )

        return redirect(url_for("dashboard"))

    return render_template("upload_listas.html")


@app.route("/", methods=["GET"])
@login_required
def dashboard():
    return render_template("dashboard.html")


# ==========================================================
#  CARTEIRINHAS – ROTA PRINCIPAL
# ==========================================================

@app.route("/carteirinhas", methods=["GET", "POST"])
@login_required
def carteirinhas():
    if request.method == "POST":
        file_path = None

        # Checkbox no form (ex.: <input type="checkbox" name="somente_com_foto">)
        somente_com_foto = request.form.get("somente_com_foto") in ("1", "on", "true", "True", "SIM", "sim")
        session["carteirinhas_somente_com_foto"] = somente_com_foto

        if "excel_file" in request.files and request.files["excel_file"].filename != "":
            file = request.files["excel_file"]
            filename = secure_filename(file.filename)
            unique_filename = f"carteirinhas_{uuid.uuid4().hex}_{filename}"
            file_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_filename)
            file.save(file_path)
            session["lista_fundamental"] = file_path
        else:
            file_path = session.get("lista_fundamental")

        if not file_path or not os.path.exists(file_path):
            flash("Nenhum arquivo selecionado. Envie a lista piloto do Fundamental.", "info")
            return redirect(url_for("carteirinhas"))

        flash("Gerando carteirinhas. Aguarde...", "info")
        html_result = gerar_html_carteirinhas(
            file_path,
            somente_com_foto=somente_com_foto,
        )
        return html_result

    # GET: opcionalmente passa o estado atual do filtro para o template marcar o checkbox
    somente_com_foto = session.get("carteirinhas_somente_com_foto", False)
    return render_template("carteirinhas.html", somente_com_foto=somente_com_foto)


#  DECLARAÇÕES – CONCLUSÃO 5º ANO (LOTE)
# ==========================================================

@app.route("/declaracao/conclusao_5ano")
@login_required
def declaracao_conclusao_5ano():
    if session.get("declaracao_tipo") != "Fundamental":
        flash(
            "As declarações em lote de 5º ano estão disponíveis apenas para o Fundamental.",
            "error",
        )
        return redirect(url_for("declaracao_tipo"))

    file_path = session.get("declaracao_excel") or session.get("lista_fundamental")

    if not file_path or not os.path.exists(file_path):
        flash(
            "Arquivo Excel do Fundamental não encontrado. "
            "Anexe a lista piloto novamente pela tela de declarações.",
            "error",
        )
        return redirect(url_for("declaracao_tipo", segmento="Fundamental"))

    planilha = pd.read_excel(file_path, sheet_name="LISTA CORRIDA")
    planilha.columns = [c.strip().upper() for c in planilha.columns]

    def format_rm(x):
        try:
            return str(int(float(x)))
        except Exception:
            return str(x)

    planilha["RM_str"] = planilha["RM"].apply(format_rm)

    registros = []

    for _, row in planilha.iterrows():
        rm_str = str(row.get("RM_str", "")).strip()
        if rm_str in ("", "0"):
            continue

        serie_raw = str(row.get("SÉRIE", "")).strip()
        if not serie_raw:
            continue

        if "5º" not in serie_raw and "5°" not in serie_raw:
            continue

        nome = str(row.get("NOME", "")).strip()
        ra = str(row.get("RA", "")).strip()

        data_nasc_val = row.get("DATA NASC.")
        if pd.notna(data_nasc_val):
            try:
                data_nasc_dt = pd.to_datetime(data_nasc_val, errors="coerce")
                data_nasc = (
                    data_nasc_dt.strftime("%d/%m/%Y")
                    if pd.notna(data_nasc_dt)
                    else "Desconhecida"
                )
            except Exception:
                data_nasc = "Desconhecida"
        else:
            data_nasc = "Desconhecida"

        horario = str(row.get("HORÁRIO", "")).strip()
        if not horario:
            horario = "Desconhecido"

        serie_fmt = serie_raw
        try:
            serie_fmt = re.sub(r"(\d+º)\s*([A-Za-z])", r"\1 ano \2", serie_fmt)
        except Exception:
            pass

        series_text = "a série subsequente"
        m = re.search(r"(\d+)º", serie_fmt)
        if m:
            try:
                next_year = int(m.group(1)) + 1
                series_text = f"{next_year}º ano"
            except Exception:
                pass

        valor_bolsa = str(row.get("BOLSA FAMILIA", "")).strip().upper()

        declaracao_text = (
            f"Declaro, para os devidos fins, que o(a) aluno(a) "
            f"<strong><u>{nome}</u></strong>, "
            f"portador(a) do RA <strong><u>{ra}</u></strong>, nascido(a) em "
            f"<strong><u>{data_nasc}</u></strong>, concluiu com êxito o "
            f"<strong><u>{serie_fmt}</u></strong>, estando apto(a) a ingressar no "
            f"<strong><u>{series_text}</u></strong>."
        )

        if valor_bolsa == "SIM":
            declaracao_text += "<br><br><strong>Observações:</strong><br>"
            declaracao_text += (
                '<label class="checkbox-label" '
                'style="display: block; text-align: justify; font-size:14px;">'
            )
            declaracao_text += (
                f'<img src="{url_for("static", filename="logos/bolsa_familia.jpg")}" '
                'alt="Bolsa Família" '
                'style="width:28px; vertical-align:middle; margin-right:5px;">'
                "O aluno é beneficiário do Programa Bolsa Família."
            )
            declaracao_text += "</label>"

        registros.append(
            {
                "nome": nome,
                "ra": ra,
                "data_nasc": data_nasc,
                "serie_fmt": serie_fmt,
                "series_text": series_text,
                "texto": declaracao_text,
            }
        )

    if not registros:
        flash("Nenhum aluno de 5º ano encontrado na lista piloto.", "error")
        return redirect(url_for("declaracao_tipo", segmento="Fundamental"))

    data_extenso_str = "Praia Grande, 22 de dezembro de 2025"
    titulo = "Declaração de Conclusão"

    registros_duas_vias = []
    for reg in registros:
        reg1 = reg.copy()
        reg1["via"] = 1
        registros_duas_vias.append(reg1)

        reg2 = reg.copy()
        reg2["via"] = 2
        registros_duas_vias.append(reg2)

    return render_template(
        "declaracao_conclusao_5ano.html",
        registros=registros_duas_vias,
        data_extenso=data_extenso_str,
        titulo=titulo,
        total=len(registros_duas_vias),
    )


# ==========================================================
#  DECLARAÇÕES – TELA ÚNICA (Fundamental / EJA / Personalizada)
# ==========================================================

@app.route("/declaracao/tipo", methods=["GET", "POST"])
@login_required
def declaracao_tipo():
    if request.method == "POST":
        modo_declaracao = request.form.get("modo_declaracao")

        # -------------------------------
        # FLUXO: DECLARAÇÃO PERSONALIZADA
        # -------------------------------
        if modo_declaracao == "personalizada":
            segmento_pers = request.form.get("segmento_personalizado")
            if segmento_pers not in ("Fundamental", "EJA"):
                flash(
                    "Selecione o segmento (Ensino Fundamental ou EJA) na declaração personalizada.",
                    "error",
                )
                return redirect(url_for("declaracao_tipo", segmento="Personalizado"))

            nome_aluno = (request.form.get("nome_aluno") or "").strip()
            data_nascimento = request.form.get("data_nascimento")
            ra = (request.form.get("ra") or "").strip()
            tipo_pers = request.form.get("tipo_declaracao_personalizada")

            if (
                not nome_aluno
                or not data_nascimento
                or not ra
                or tipo_pers not in ("Conclusao", "MatriculaCancelada", "NCOM")
            ):
                flash(
                    "Preencha todos os dados obrigatórios da declaração personalizada.",
                    "error",
                )
                return redirect(url_for("declaracao_tipo", segmento="Personalizado"))

            dados_personalizados = {
                "segmento": segmento_pers,
                "nome_aluno": nome_aluno,
                "data_nascimento": data_nascimento,
                "ra": ra,
                "tipo_declaracao": tipo_pers,
            }

            if tipo_pers == "Conclusao":
                ano_serie_concluida = (request.form.get("ano_serie_concluida") or "").strip()
                ano_conclusao = (request.form.get("ano_conclusao") or "").strip()
                deve_hist_unidade = request.form.get("deve_historico_unidade")
                semestre_conclusao = (request.form.get("semestre_conclusao") or "").strip()

                campos_invalidos = (
                    not ano_serie_concluida
                    or not ano_conclusao
                    or deve_hist_unidade not in ("Sim", "Não")
                )

                if segmento_pers == "EJA" and not semestre_conclusao:
                    campos_invalidos = True

                if campos_invalidos:
                    flash(
                        "Preencha todos os campos da declaração personalizada de conclusão "
                        "(para EJA é obrigatório informar o semestre).",
                        "error",
                    )
                    return redirect(url_for("declaracao_tipo", segmento="Personalizado"))

                dados_personalizados.update(
                    {
                        "ano_serie_concluida": ano_serie_concluida,
                        "ano_conclusao": ano_conclusao,
                        "deve_historico_unidade": (deve_hist_unidade == "Sim"),
                        "semestre_conclusao": semestre_conclusao,
                    }
                )

            elif tipo_pers == "MatriculaCancelada":
                ano_serie_matricula = (request.form.get("ano_serie_matricula") or "").strip()
                ano_matricula = (request.form.get("ano_matricula") or "").strip()
                semestre_matricula = (request.form.get("semestre_matricula") or "").strip()

                if not ano_serie_matricula or not ano_matricula or not semestre_matricula:
                    flash(
                        "Preencha todos os campos da declaração de matrícula cancelada.",
                        "error",
                    )
                    return redirect(url_for("declaracao_tipo", segmento="Personalizado"))

                dados_personalizados.update(
                    {
                        "ano_serie_matricula": ano_serie_matricula,
                        "ano_matricula": ano_matricula,
                        "semestre_matricula": semestre_matricula,
                    }
                )

            elif tipo_pers == "NCOM":
                ano_serie_vaga = (request.form.get("ano_serie_vaga") or "").strip()
                ano_referencia_ncom = (request.form.get("ano_referencia_ncom") or "").strip()
                semestre_referencia_ncom = (
                    request.form.get("semestre_referencia_ncom") or ""
                ).strip()

                if not ano_serie_vaga or not ano_referencia_ncom:
                    flash(
                        "Preencha todos os campos obrigatórios da declaração de Não Comparecimento (NCOM).",
                        "error",
                    )
                    return redirect(url_for("declaracao_tipo", segmento="Personalizado"))

                dados_personalizados.update(
                    {
                        "ano_serie_vaga": ano_serie_vaga,
                        "ano_referencia_ncom": ano_referencia_ncom,
                        "semestre_referencia_ncom": semestre_referencia_ncom,
                    }
                )

            declaracao_html = gerar_declaracao_personalizada(dados_personalizados)

            if declaracao_html is None:
                flash(
                    "Não foi possível gerar a declaração personalizada. Verifique os dados informados.",
                    "error",
                )
                return redirect(url_for("declaracao_tipo", segmento="Personalizado"))

            return declaracao_html

        # -------------------------------
        # FLUXO NORMAL: FUNDAMENTAL / EJA
        # -------------------------------
        segmento = request.form.get("segmento_escolhido")
        if segmento not in ("Fundamental", "EJA"):
            flash("Selecione se a declaração é do Fundamental ou EJA antes de gerar.", "error")
            return redirect(url_for("declaracao_tipo"))

        rm = (request.form.get("rm") or "").strip()
        tipo = (request.form.get("tipo") or "").strip()

        tipo_lower = tipo.lower()
        if tipo_lower in ("transferencia", "transferência"):
            tipo = "Transferencia"
        elif tipo_lower in ("conclusao", "conclusão"):
            tipo = "Conclusão"
        elif tipo_lower in ("frequencia", "frequência"):
            tipo = "Frequencia"

        deve_historico_str = request.form.get("deve_historico")

        unidade_select = (request.form.get("unidade_anterior_select") or "").strip()
        unidade_manual = (request.form.get("unidade_anterior_manual") or "").strip()
        unidade_anterior = unidade_select or unidade_manual

        file_path = None
        excel_file = request.files.get("excel_file")
        novo_upload = excel_file is not None and excel_file.filename

        if novo_upload:
            filename = secure_filename(excel_file.filename)
            unique_filename = f"declaracao_{uuid.uuid4().hex}_{filename}"
            file_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_filename)
            excel_file.save(file_path)

            if segmento == "Fundamental":
                session["lista_fundamental"] = file_path
            else:
                session["lista_eja"] = file_path
        else:
            if segmento == "Fundamental":
                file_path = session.get("lista_fundamental")
            else:
                file_path = session.get("lista_eja")

        if not file_path or not os.path.exists(file_path):
            flash(
                "Nenhuma lista piloto encontrada para este segmento. Anexe o arquivo em Excel.",
                "error",
            )
            return redirect(url_for("declaracao_tipo", segmento=segmento))

        session["declaracao_tipo"] = segmento
        session["declaracao_excel"] = file_path

        if novo_upload and (not rm or not tipo):
            flash(
                "Lista piloto carregada com sucesso. Agora selecione o aluno e o tipo de declaração.",
                "success",
            )
            return redirect(url_for("declaracao_tipo", segmento=segmento))

        if not rm or not tipo:
            flash("Escolha o aluno e o tipo de declaração.", "error")
            return redirect(url_for("declaracao_tipo", segmento=segmento))

        if tipo in ("Transferencia", "Conclusão"):
            if deve_historico_str not in ("sim", "nao"):
                flash("Por favor, responda se o aluno deve o histórico escolar.", "error")
                return redirect(url_for("declaracao_tipo", segmento=segmento))

            if deve_historico_str == "sim" and not unidade_anterior:
                flash(
                    "Informe a unidade escolar anterior para a qual o aluno deve o histórico.",
                    "error",
                )
                return redirect(url_for("declaracao_tipo", segmento=segmento))

            deve_historico = deve_historico_str == "sim"
        else:
            deve_historico = False
            unidade_anterior = ""

        dados_frequencia = None

        if tipo == "Frequencia":
            meses = [
                ("jan", "Janeiro"),
                ("fev", "Fevereiro"),
                ("mar", "Março"),
                ("abr", "Abril"),
                ("mai", "Maio"),
                ("jun", "Junho"),
                ("jul", "Julho"),
                ("ago", "Agosto"),
                ("set", "Setembro"),
                ("out", "Outubro"),
                ("nov", "Novembro"),
                ("dez", "Dezembro"),
            ]

            dados_frequencia = {"meses": []}
            algum_valido = False

            for mes_id, mes_nome in meses:
                dias_raw = (request.form.get(f"freq_{mes_id}_dias") or "").strip()
                faltas_raw = (request.form.get(f"freq_{mes_id}_faltas") or "").strip()

                if not dias_raw and not faltas_raw:
                    dados_frequencia["meses"].append(
                        {
                            "id": mes_id,
                            "nome": mes_nome,
                            "dias_letivos": None,
                            "faltas": None,
                            "frequencia": None,
                            "preenchido": False,
                        }
                    )
                    continue

                try:
                    dias = float(dias_raw.replace(",", ".")) if dias_raw else None
                    faltas = float(faltas_raw.replace(",", ".")) if faltas_raw else None
                except ValueError:
                    flash(
                        "Verifique os valores de dias letivos e faltas informados na frequência.",
                        "error",
                    )
                    return redirect(url_for("declaracao_tipo", segmento=segmento))

                if dias is None or faltas is None:
                    flash(
                        "Para cada mês de frequência preenchido, informe tanto os dias letivos quanto as faltas.",
                        "error",
                    )
                    return redirect(url_for("declaracao_tipo", segmento=segmento))

                if dias <= 0 or faltas < 0 or faltas > dias:
                    flash(
                        "Os valores de dias letivos e faltas são inválidos em um ou mais meses. "
                        "Verifique e tente novamente.",
                        "error",
                    )
                    return redirect(url_for("declaracao_tipo", segmento=segmento))

                freq_percent = ((dias - faltas) / dias) * 100.0
                algum_valido = True

                dados_frequencia["meses"].append(
                    {
                        "id": mes_id,
                        "nome": mes_nome,
                        "dias_letivos": dias,
                        "faltas": faltas,
                        "frequencia": round(freq_percent, 1),
                        "preenchido": True,
                    }
                )

            if not algum_valido:
                flash(
                    "Informe ao menos um mês de frequência com dias letivos e faltas válidos.",
                    "error",
                )
                return redirect(url_for("declaracao_tipo", segmento=segmento))

        declaracao_html = gerar_declaracao_escolar(
            file_path=file_path,
            rm=rm,
            tipo=tipo,
            deve_historico=deve_historico,
            unidade_anterior=unidade_anterior,
            dados_frequencia=dados_frequencia,
        )

        if declaracao_html is None:
            flash("Aluno não encontrado na lista piloto.", "error")
            return redirect(url_for("declaracao_tipo", segmento=segmento))

        return declaracao_html

    # --------------------------------------
    # GET: EXIBE A TELA
    # --------------------------------------
    segmento = request.args.get("segmento")
    if segmento not in ("Fundamental", "EJA", "Personalizado"):
        segmento = None

    alunos = []
    tem_lista = False

    if segmento == "Fundamental":
        file_path = session.get("lista_fundamental")
        if file_path and os.path.exists(file_path):
            tem_lista = True
            session["declaracao_tipo"] = "Fundamental"
            session["declaracao_excel"] = file_path

            planilha = pd.read_excel(file_path, sheet_name="LISTA CORRIDA")

            def format_rm(x):
                try:
                    return str(int(float(x)))
                except Exception:
                    return str(x)

            planilha["RM_str"] = planilha["RM"].apply(format_rm)
            alunos_df = (
                planilha[planilha["RM_str"] != "0"][["RM_str", "NOME"]].drop_duplicates()
            )

            alunos = [
                {"rm": row["RM_str"], "nome": row["NOME"]}
                for _, row in alunos_df.iterrows()
            ]

    elif segmento == "EJA":
        # EJA permanece suportada, porém agora é opcional no sistema como um todo.
        file_path = session.get("lista_eja")
        if file_path and os.path.exists(file_path):
            tem_lista = True
            session["declaracao_tipo"] = "EJA"
            session["declaracao_excel"] = file_path

            df = pd.read_excel(file_path, sheet_name=0, header=None, skiprows=1)
            df["RM_str"] = df.iloc[:, 2].apply(
                lambda x: str(int(x)) if pd.notna(x) and float(x) != 0 else ""
            )
            df["NOME"] = df.iloc[:, 3]
            alunos_df = df[df["RM_str"] != ""][["RM_str", "NOME"]].drop_duplicates()

            alunos = [
                {"rm": row["RM_str"], "nome": row["NOME"]}
                for _, row in alunos_df.iterrows()
            ]
        else:
            tem_lista = False
            alunos = []

    dashboard_url = url_for("dashboard")
    conclusao_5ano_url = url_for("declaracao_conclusao_5ano")
    escolaridade_5ano_url = url_for("declaracao_escolaridade_5ano")

    return render_template(
        "declaracao_tipo.html",
        segmento=segmento,
        tem_lista=tem_lista,
        alunos=alunos,
        dashboard_url=dashboard_url,
        conclusao_5ano_url=conclusao_5ano_url,
        escolaridade_5ano_url=escolaridade_5ano_url,
    )


@app.route("/declaracao/escolaridade_5ano")
@login_required
def declaracao_escolaridade_5ano(file_path_arg=None):
    if file_path_arg:
        file_path = file_path_arg
    else:
        file_path = session.get("declaracao_excel")

    if session.get("declaracao_tipo") != "Fundamental":
        flash(
            "As declarações de escolaridade de 5º ano só podem ser geradas "
            "com a lista piloto do Ensino Fundamental.",
            "error",
        )
        return redirect(url_for("declaracao_tipo", segmento="Fundamental"))

    if not file_path or not os.path.exists(file_path):
        flash(
            "Nenhuma lista piloto do Ensino Fundamental está carregada. "
            "Anexe a lista piloto novamente para gerar as declarações.",
            "error",
        )
        return redirect(url_for("declaracao_tipo", segmento="Fundamental"))

    registros, data_extenso_str, titulo = gerar_lote_escolaridade_5ano(file_path)

    if not registros:
        flash(
            "Nenhum aluno de 5º ano foi encontrado na lista piloto para "
            "gerar as declarações de escolaridade.",
            "error",
        )
        return redirect(url_for("declaracao_tipo", segmento="Fundamental"))

    return render_template(
        "declaracao_escolaridade_5ano.html",
        registros=registros,
        data_extenso=data_extenso_str,
        titulo=titulo,
    )


# ==========================================================
#  UPLOAD DE FOTOS (CARTEIRINHAS)
# ==========================================================

@app.route("/upload_foto", methods=["POST"])
@login_required
def upload_foto():
    rm = (request.form.get("rm") or "").strip()
    if not rm:
        flash("RM não fornecido.", "error")
        return redirect(url_for("carteirinhas"))

    file = request.files.get("foto_file")
    if not file or file.filename == "":
        flash("Nenhuma foto selecionada.", "error")
        return redirect(url_for("carteirinhas"))

    if not allowed_file(file.filename):
        flash("Formato de imagem não permitido. Envie JPG, PNG, GIF ou BMP.", "error")
        return redirect(url_for("carteirinhas"))

    original_filename = secure_filename(file.filename)
    _, ext = os.path.splitext(original_filename)

    fotos_dir = os.path.join("static", "fotos")
    os.makedirs(fotos_dir, exist_ok=True)

    new_filename = secure_filename(f"{rm}{ext.lower()}")
    file_path = os.path.join(fotos_dir, new_filename)
    file.save(file_path)

    flash("Foto anexada com sucesso.", "success")
    return redirect(url_for("carteirinhas"))


@app.route("/upload_multiplas_fotos", methods=["POST"])
@login_required
def upload_multiplas_fotos():
    rms = request.form.getlist("rm[]")
    files = request.files.getlist("foto_file[]")

    if not files:
        flash("Nenhuma foto enviada.", "error")
        return redirect(url_for("carteirinhas"))

    fotos_dir = os.path.join("static", "fotos")
    os.makedirs(fotos_dir, exist_ok=True)

    total_salvas = 0
    total_ignoradas = 0

    for rm, file in zip(rms, files):
        rm = (rm or "").strip()

        if not rm or not file or file.filename == "" or not allowed_file(file.filename):
            total_ignoradas += 1
            continue

        original_filename = secure_filename(file.filename)
        _, ext = os.path.splitext(original_filename)
        new_filename = secure_filename(f"{rm}{ext.lower()}")
        file_path = os.path.join(fotos_dir, new_filename)
        file.save(file_path)
        total_salvas += 1

    if total_salvas:
        msg = f"Foto(s) anexada(s) com sucesso: {total_salvas} arquivo(s)."
        if total_ignoradas:
            msg += f" {total_ignoradas} arquivo(s) foram ignorados por RM ou formato inválido."
        flash(msg, "success")
    else:
        flash("Nenhuma foto válida foi enviada.", "error")

    return redirect(url_for("carteirinhas"))


@app.route("/upload_inline_foto", methods=["POST"])
@login_required
def upload_inline_foto():
    file = request.files.get("foto_file")
    rm = (request.form.get("rm") or "").strip()

    if not rm:
        return jsonify({"error": "RM não fornecido"}), 400

    if not file or file.filename == "":
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    if not allowed_file(file.filename):
        return jsonify({"error": "Formato de imagem não permitido"}), 400

    original_filename = secure_filename(file.filename)
    _, ext = os.path.splitext(original_filename)

    fotos_dir = os.path.join("static", "fotos")
    os.makedirs(fotos_dir, exist_ok=True)

    new_filename = secure_filename(f"{rm}{ext.lower()}")
    file_path = os.path.join(fotos_dir, new_filename)
    file.save(file_path)

    return jsonify({"url": f"/static/fotos/{new_filename}", "message": "Foto anexada com sucesso"}), 200


# ==========================================================
#  QUADROS – HELPERS COMUNS + EJA (LIGA/DESLIGA)
#  (FIX RÁPIDO E SEGURO: sem colisão de nomes, sem sobrescrita de helpers)
# ==========================================================

import os
import re
import uuid
import copy
import string
import unicodedata
from contextlib import contextmanager
from datetime import datetime
from io import BytesIO
from typing import Optional, Iterable, Tuple, Dict

import pandas as pd
from flask import (
    request,
    redirect,
    url_for,
    render_template,
    flash,
    session,
    current_app,
    send_file,
)
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.cell import MergedCell

# ----------------------------------------------------------
# EJA: liga/desliga (FIX)
# ----------------------------------------------------------
def _is_eja_enabled() -> bool:
    """
    Desativada por padrão:
      - Ative via variável de ambiente: ENABLE_EJA=1
      - Ou via config: app.config["ENABLE_EJA"] = True

    FIX: remove duplicidade/override incorreto.
    """
    def _to_bool(v) -> bool:
        if v is None:
            return False
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        return s in ("1", "true", "t", "yes", "y", "on")

    # 1) prioridade para config do Flask
    cfg = None
    try:
        cfg = current_app.config.get("ENABLE_EJA", None)
    except Exception:
        cfg = None

    if cfg is not None:
        return _to_bool(cfg)

    # 2) fallback: env var
    return _to_bool(os.getenv("ENABLE_EJA", "0"))


def _save_upload_to_session(file_storage, session_key: str, prefix: str) -> str:
    """
    Salva arquivo enviado em UPLOAD_FOLDER e grava em session[session_key].
    Retorna o path salvo.
    """
    filename = secure_filename(file_storage.filename)
    unique_filename = f"{prefix}_{uuid.uuid4().hex}_{filename}"

    upload_folder = None
    try:
        upload_folder = current_app.config.get("UPLOAD_FOLDER")
    except Exception:
        upload_folder = None

    if not upload_folder:
        # mantém compatibilidade com seu app.py original (que define app.config["UPLOAD_FOLDER"])
        upload_folder = "uploads"

    file_path = os.path.join(upload_folder, unique_filename)
    file_storage.save(file_path)
    session[session_key] = file_path
    return file_path


def _find_sheet_case_insensitive(wb, target_name: str):
    """
    Busca uma aba ignorando maiúsculas/minúsculas e espaços.
    Retorna o nome real encontrado ou None.
    """
    target = (target_name or "").strip().lower()
    for name in wb.sheetnames:
        if (name or "").strip().lower() == target:
            return name
    return None


# ----------------------------------------------------------
# Helper comum: escrita segura em célula mesclada
# ----------------------------------------------------------
def set_merged_cell_value(ws, cell_coord: str, value):
    """
    Atualiza o valor de uma célula (inclusive se estiver mesclada),
    preservando a mesclagem.
    """
    cell = ws[cell_coord]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell_coord in merged_range:
                range_str = str(merged_range)
                ws.unmerge_cells(range_str)
                min_col, min_row, _, _ = merged_range.bounds
                top_left = ws.cell(row=min_row, column=min_col)
                top_left.value = value
                ws.merge_cells(range_str)
                return
    ws[cell_coord].value = value


# ----------------------------------------------------------
# Helpers comuns: normalização de cabeçalho e strings
# ----------------------------------------------------------
def _safe_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    return str(v).strip()


def _norm_header_compact(text: str) -> str:
    """
    Normaliza cabeçalho: remove acentos e tudo que não for A-Z0-9.
    Fica robusto para variações: 'LOCAL TE', 'LOCAL_TE', 'local te', etc.
    """
    if text is None:
        return ""
    s = str(text).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper()
    s = re.sub(r"[^A-Z0-9]+", "", s)
    return s


def _build_colmap(df: pd.DataFrame) -> dict:
    """Mapeia cabeçalho normalizado -> nome real da coluna (primeira ocorrência)."""
    m = {}
    for col in df.columns:
        k = _norm_header_compact(col)
        if k and k not in m:
            m[k] = col
    return m


def _pick_col(colmap: dict, *candidates: str):
    """Retorna o nome real da coluna a partir de candidatos (por normalização compacta)."""
    for cand in candidates:
        k = _norm_header_compact(cand)
        if k in colmap:
            return colmap[k]
    return None


def _find_df_col(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    """
    Encontra uma coluna do DataFrame por lista de candidatos (nome),
    comparando com normalização compacta.
    """
    if df is None or df.empty:
        return None
    colmap = _build_colmap(df)
    return _pick_col(colmap, *list(candidates))


def _is_missing_value(val) -> bool:
    """Considera vazio/0/-/nan como ausente."""
    s = _safe_str(val)
    if not s:
        return True
    return s.lower() in {"0", "-", "nan", "none", "null"}


@contextmanager
def _temp_unprotect_sheet(ws):
    """Desabilita proteção da planilha temporariamente para escrita e restaura depois."""
    original = copy.copy(ws.protection)
    try:
        ws.protection.sheet = False
        yield
    finally:
        ws.protection = original


# ==========================================================
#  QUADROS – MENU PRINCIPAL
# ==========================================================
@app.route("/quadros")
@login_required
def quadros():
    return render_template("quadros.html")


# ==========================================================
#  QUADRO – INCLUSÃO (DESATIVADO / ARQUIVADO)
# ==========================================================
@app.route("/quadros/inclusao", methods=["GET", "POST"])
@login_required
def quadros_inclusao():
    flash("Quadro de Inclusão foi migrado para Drive e está desativado no sistema no momento.", "info")
    return redirect(url_for("quadros"))


# ==========================================================
#  QUADRO – ATENDIMENTO MENSAL (CORRIGIDO / FUTURE-PROOF)
# ==========================================================

TURMAS_MAX = list(string.ascii_uppercase[:14])  # A..N

ATENDIMENTO_CONFIG = {
    "MODEL_BLOCKS": {
        # 1º ano SEMPRE zerado: Masc + Fem + Total
        "1º": {"start_row": 19, "masc_col": "B", "fem_col": "C", "total_col": "D"},
        "2º": {"start_row": 37, "masc_col": "B", "fem_col": "C", "total_col": "D"},
        "3º": {"start_row": 55, "masc_col": "B", "fem_col": "C", "total_col": "D"},
        "4º": {"start_row": 73, "masc_col": "B", "fem_col": "C", "total_col": "D"},
        "5º": {"start_row": 91, "masc_col": "B", "fem_col": "C", "total_col": "D"},
    },

    # Aba "Total de Alunos" (layout padrão 2026)
    "PILOTO_COLS_DEFAULT": {
        "serie_col": 3,    # C
        "turma_col": 4,    # D
        "ma_masc_col": 7,  # G
        "ma_fem_col": 8,   # H
        "ma_total_col": 9  # I
    },

    # Fallback por blocos (Prioridade 2)
    "FALLBACK_START": {
        "2º": {"row": 6,  "masc_col": 7, "fem_col": 8},
        "3º": {"row": 14, "masc_col": 7, "fem_col": 8},
        "4º": {"row": 21, "masc_col": 7, "fem_col": 8},
        "5º": {"row": 29, "masc_col": 7, "fem_col": 8},
    },

    "TOTALS": {
        "manha_total_cell": (38, 9),  # I38
        "tarde_total_cell": (40, 9),  # I40
        "modelo_manha_addr": "R20",
        "modelo_tarde_addr": "R28",
    },

    "ENABLE_DEBUG_LOG": True,
}


def _safe_int(val, default=0):
    if val is None or val == "":
        return default
    if isinstance(val, bool):
        return int(val)
    if isinstance(val, (int, float)):
        try:
            return int(val)
        except Exception:
            return default
    try:
        s = str(val).strip()
        s = s.replace(".", "").replace(",", ".")
        return int(float(s))
    except Exception:
        return default


def _norm_serie(val):
    if val is None:
        return None
    s = str(val).strip()
    m = re.search(r"(\d)\s*[ºo°]?", s, flags=re.IGNORECASE)
    if not m:
        return None
    n = m.group(1)
    if n in {"1", "2", "3", "4", "5"}:
        return f"{n}º"
    return None


def _extract_turma_letter(val):
    if val is None:
        return None
    s = str(val).strip().upper()
    m = re.search(r"\b([A-N])\b", s)
    if m:
        return m.group(1)
    m2 = re.search(r"([A-N])", s)
    return m2.group(1) if m2 else None


def _condense_letters(letters):
    if not letters:
        return "-"
    idxs = sorted({TURMAS_MAX.index(x) for x in letters if x in TURMAS_MAX})
    out = []
    i = 0
    while i < len(idxs):
        start = idxs[i]
        j = i
        while j + 1 < len(idxs) and idxs[j + 1] == idxs[j] + 1:
            j += 1
        if j == i:
            out.append(TURMAS_MAX[start])
        else:
            out.append(f"{TURMAS_MAX[start]}-{TURMAS_MAX[idxs[j]]}")
        i = j + 1
    return ", ".join(out)


def _detect_ma_columns(ws_total):
    cfg = ATENDIMENTO_CONFIG["PILOTO_COLS_DEFAULT"].copy()
    try:
        for r in range(1, 20):
            for c in range(1, 30):
                v = ws_total.cell(row=r, column=c).value
                if v and "MATRICULAS" in str(v).upper():
                    ma_col = c
                    cfg["ma_masc_col"] = ma_col
                    cfg["ma_fem_col"] = ma_col + 1
                    cfg["ma_total_col"] = ma_col + 2
                    return cfg
    except Exception:
        pass
    return cfg


def _extract_by_cols(ws_total, serie_label, debug_log):
    cols = _detect_ma_columns(ws_total)
    sc = cols["serie_col"]
    tc = cols["turma_col"]
    mc = cols["ma_masc_col"]
    fc = cols["ma_fem_col"]

    found = {}
    duplicates = []

    max_row = ws_total.max_row or 0
    limit = min(max_row, 300)

    for r in range(1, limit + 1):
        serie = _norm_serie(ws_total.cell(row=r, column=sc).value)
        if serie != serie_label:
            continue

        turma = _extract_turma_letter(ws_total.cell(row=r, column=tc).value)
        if not turma:
            continue

        masc = _safe_int(ws_total.cell(row=r, column=mc).value, 0)
        fem = _safe_int(ws_total.cell(row=r, column=fc).value, 0)

        if turma in found:
            duplicates.append(turma)
            found[turma] = (found[turma][0] + masc, found[turma][1] + fem)
        else:
            found[turma] = (masc, fem)

    if duplicates:
        debug_log.append(f"[{serie_label}] AVISO: turmas duplicadas (somadas): {sorted(set(duplicates))}")

    return found


def _extract_by_fallback_block(ws_total, serie_label, debug_log):
    fb = ATENDIMENTO_CONFIG["FALLBACK_START"].get(serie_label)
    if not fb:
        return {}

    serie_col = ATENDIMENTO_CONFIG["PILOTO_COLS_DEFAULT"]["serie_col"]  # C
    turma_col = ATENDIMENTO_CONFIG["PILOTO_COLS_DEFAULT"]["turma_col"]  # D

    row = fb["row"]
    mc = fb["masc_col"]
    fc = fb["fem_col"]

    found = {}
    for _ in range(len(TURMAS_MAX)):
        serie_here = _norm_serie(ws_total.cell(row=row, column=serie_col).value)
        if serie_here != serie_label:
            break

        turma = _extract_turma_letter(ws_total.cell(row=row, column=turma_col).value)
        if not turma:
            break

        masc = _safe_int(ws_total.cell(row=row, column=mc).value, 0)
        fem = _safe_int(ws_total.cell(row=row, column=fc).value, 0)
        found[turma] = (masc, fem)

        row += 1

    debug_log.append(f"[{serie_label}] fallback usado: capturadas {len(found)} turmas (parada por mudança de série na col C).")
    return found


def _write_block(ws_modelo, serie_label, turma_data, debug_log):
    block = ATENDIMENTO_CONFIG["MODEL_BLOCKS"][serie_label]
    start = block["start_row"]
    masc_col = block["masc_col"]
    fem_col = block["fem_col"]
    total_col = block["total_col"]

    found_letters = [t for t in TURMAS_MAX if t in turma_data]
    missing_letters = [t for t in TURMAS_MAX if t not in turma_data]

    for i, turma in enumerate(TURMAS_MAX):
        r = start + i

        masc = turma_data.get(turma, (0, 0))[0]
        fem = turma_data.get(turma, (0, 0))[1]

        set_merged_cell_value(ws_modelo, f"{masc_col}{r}", masc)

        if fem_col:
            set_merged_cell_value(ws_modelo, f"{fem_col}{r}", fem)

        if total_col and fem_col:
            # Força fórmula para evitar valor residual no template
            set_merged_cell_value(ws_modelo, f"{total_col}{r}", f"={masc_col}{r}+{fem_col}{r}")

    debug_log.append(f"[{serie_label}] preenchido: {_condense_letters(found_letters)}; zerado: {_condense_letters(missing_letters)}")


def _read_total(ws_total, row, col, debug_log, label):
    v = ws_total.cell(row=row, column=col).value
    out = _safe_int(v, 0)
    debug_log.append(f"[TOTAL] {label}: lido {out} de ({row},{col}).")
    return out


@app.route('/quadros/atendimento_mensal', methods=['GET', 'POST'])
@login_required
def quadro_atendimento_mensal():
    if request.method == 'POST':
        debug_log = []
        enable_eja = _is_eja_enabled()

        responsavel = (request.form.get("responsavel") or "").strip()
        rf = (request.form.get("rf") or "").strip()

        mes_ref_raw = (request.form.get("mes_ref") or "").strip()
        mes_ref = None
        if mes_ref_raw:
            m = re.match(r"^\s*(\d{4})-(\d{2})\s*$", mes_ref_raw)  # YYYY-MM
            if m:
                mes_ref = f"{m.group(2)}/{m.group(1)}"
            else:
                m2 = re.match(r"^\s*(\d{2})/(\d{4})\s*$", mes_ref_raw)  # MM/YYYY
                if m2:
                    mes_ref = f"{m2.group(1)}/{m2.group(2)}"
        if not mes_ref:
            mes_ref = datetime.now().strftime("%m/%Y")

        fundamental_file = request.files.get('lista_fundamental')
        eja_file = request.files.get('lista_eja')

        if fundamental_file and fundamental_file.filename != '':
            _save_upload_to_session(fundamental_file, 'lista_fundamental', prefix='atendimento')

        if enable_eja and eja_file and eja_file.filename != '':
            _save_upload_to_session(eja_file, 'lista_eja', prefix='atendimento_eja')

        file_path = session.get('lista_fundamental')
        if not file_path or not os.path.exists(file_path):
            flash("Nenhum arquivo da Lista Piloto FUNDAMENTAL disponível.", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        model_path = os.path.join("modelos", "Quadro de Atendimento Mensal - Modelo.xlsx")
        if not os.path.exists(model_path):
            flash("Modelo Atendimento Mensal não encontrado.", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        try:
            with open(model_path, "rb") as f:
                wb_modelo = load_workbook(f, data_only=False)
        except Exception as e:
            flash(f"Erro ao ler o modelo de atendimento mensal: {str(e)}", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        ws_modelo = wb_modelo.worksheets[1] if len(wb_modelo.worksheets) > 1 else wb_modelo.active

        set_merged_cell_value(ws_modelo, "B5", "E.M José Padin Mouta")
        set_merged_cell_value(ws_modelo, "C6", responsavel or "-")
        set_merged_cell_value(ws_modelo, "B7", rf or "-")
        set_merged_cell_value(ws_modelo, "A13", mes_ref)
        debug_log.append(f"[HEADER] responsavel='{responsavel or '-'}' rf='{rf or '-'}' mes_ref='{mes_ref}'")

        try:
            wb_lista = load_workbook(file_path, data_only=True, read_only=True)
        except Exception:
            flash("Erro ao ler o arquivo da Lista Piloto FUNDAMENTAL.", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        sheet_name = _find_sheet_case_insensitive(wb_lista, "Total de Alunos")
        if not sheet_name:
            flash("A aba 'Total de Alunos' não foi encontrada na Lista Piloto FUNDAMENTAL.", "error")
            return redirect(url_for('quadro_atendimento_mensal'))

        ws_total = wb_lista[sheet_name]

        # 1º ano: SEMPRE ZERADO (Masc + Fem + Total)
        _write_block(ws_modelo, "1º", {}, debug_log)

        # séries 2º..5º
        for serie_label in ["2º", "3º", "4º", "5º"]:
            data = _extract_by_cols(ws_total, serie_label, debug_log)
            if not data:
                debug_log.append(f"[{serie_label}] prioridade 1 (C/D) não encontrou turmas; tentando fallback...")
                data = _extract_by_fallback_block(ws_total, serie_label, debug_log)
            _write_block(ws_modelo, serie_label, data, debug_log)

        # totais manhã / tarde
        tcfg = ATENDIMENTO_CONFIG["TOTALS"]
        manha = _read_total(ws_total, *tcfg["manha_total_cell"], debug_log=debug_log, label="Manhã")
        tarde = _read_total(ws_total, *tcfg["tarde_total_cell"], debug_log=debug_log, label="Tarde")

        set_merged_cell_value(ws_modelo, tcfg["modelo_manha_addr"], manha)
        set_merged_cell_value(ws_modelo, "R24", "-")
        set_merged_cell_value(ws_modelo, tcfg["modelo_tarde_addr"], tarde)

        # EJA: mantém comportamento atual
        def _preencher_eja_zerado():
            cells_zero = [
                "L19", "L20", "L21", "L22",
                "M19", "M20", "M21", "M22",
                "L27", "L28", "L29", "L30",
                "M27", "M28", "M29", "M30",
                "L35", "L36", "L37",
                "M35", "M36", "M37",
                "R32",
            ]
            for addr in cells_zero:
                set_merged_cell_value(ws_modelo, addr, 0)
            set_merged_cell_value(ws_modelo, "R24", "-")

        if not enable_eja:
            _preencher_eja_zerado()
            debug_log.append("[EJA] desativada: bloco EJA zerado.")
        else:
            eja_path = session.get('lista_eja')
            if not eja_path or not os.path.exists(eja_path):
                _preencher_eja_zerado()
                debug_log.append("[EJA] habilitada, mas sem arquivo: bloco EJA zerado.")
            else:
                try:
                    wb_eja = load_workbook(eja_path, data_only=True, read_only=True)
                except Exception as e:
                    flash(f"Erro ao ler a Lista Piloto EJA: {str(e)}. Gerando sem EJA.", "warning")
                    _preencher_eja_zerado()
                    debug_log.append(f"[EJA] erro ao ler: {e}. Bloco EJA zerado.")
                else:
                    sheet_name_eja = _find_sheet_case_insensitive(wb_eja, "Total de Alunos")
                    if not sheet_name_eja:
                        flash("A aba 'Total de Alunos' não foi encontrada na Lista Piloto EJA. Gerando sem EJA.", "warning")
                        _preencher_eja_zerado()
                        debug_log.append("[EJA] aba Total de Alunos não encontrada. Bloco EJA zerado.")
                    else:
                        ws_total_eja = wb_eja[sheet_name_eja]
                        set_merged_cell_value(ws_modelo, "L19", ws_total_eja.cell(row=6, column=5).value)
                        set_merged_cell_value(ws_modelo, "L20", ws_total_eja.cell(row=7, column=5).value)
                        set_merged_cell_value(ws_modelo, "L21", ws_total_eja.cell(row=8, column=5).value)
                        set_merged_cell_value(ws_modelo, "L22", ws_total_eja.cell(row=9, column=5).value)

                        set_merged_cell_value(ws_modelo, "M19", ws_total_eja.cell(row=6, column=6).value)
                        set_merged_cell_value(ws_modelo, "M20", ws_total_eja.cell(row=7, column=6).value)
                        set_merged_cell_value(ws_modelo, "M21", ws_total_eja.cell(row=8, column=6).value)
                        set_merged_cell_value(ws_modelo, "M22", ws_total_eja.cell(row=9, column=6).value)

                        set_merged_cell_value(ws_modelo, "L27", ws_total_eja.cell(row=11, column=5).value)
                        set_merged_cell_value(ws_modelo, "L28", ws_total_eja.cell(row=12, column=5).value)
                        set_merged_cell_value(ws_modelo, "L29", ws_total_eja.cell(row=13, column=5).value)
                        set_merged_cell_value(ws_modelo, "L30", ws_total_eja.cell(row=14, column=5).value)

                        set_merged_cell_value(ws_modelo, "M27", ws_total_eja.cell(row=11, column=6).value)
                        set_merged_cell_value(ws_modelo, "M28", ws_total_eja.cell(row=12, column=6).value)
                        set_merged_cell_value(ws_modelo, "M29", ws_total_eja.cell(row=13, column=6).value)
                        set_merged_cell_value(ws_modelo, "M30", ws_total_eja.cell(row=14, column=6).value)

                        set_merged_cell_value(ws_modelo, "L35", ws_total_eja.cell(row=16, column=5).value)
                        set_merged_cell_value(ws_modelo, "L36", ws_total_eja.cell(row=17, column=5).value)
                        set_merged_cell_value(ws_modelo, "L37", ws_total_eja.cell(row=18, column=5).value)

                        set_merged_cell_value(ws_modelo, "M35", ws_total_eja.cell(row=16, column=6).value)
                        set_merged_cell_value(ws_modelo, "M36", ws_total_eja.cell(row=17, column=6).value)
                        set_merged_cell_value(ws_modelo, "M37", ws_total_eja.cell(row=18, column=6).value)

                        set_merged_cell_value(ws_modelo, "R32", ws_total_eja.cell(row=20, column=7).value)
                        set_merged_cell_value(ws_modelo, "R24", "-")
                        debug_log.append("[EJA] preenchida com sucesso.")

        if ATENDIMENTO_CONFIG["ENABLE_DEBUG_LOG"]:
            try:
                current_app.logger.info("=== DEBUG ATENDIMENTO MENSAL ===\n%s", "\n".join(debug_log))
            except Exception:
                print("=== DEBUG ATENDIMENTO MENSAL ===")
                print("\n".join(debug_log))

        output = BytesIO()
        wb_modelo.save(output)
        output.seek(0)

        filename = f"Quadro de Atendimento Mensal - {datetime.now().strftime('%d%m')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return render_template('quadro_atendimento_mensal.html', enable_eja=_is_eja_enabled())


# ==========================================================
#  QUADRO – TRANSFERÊNCIAS (FIX H/I/J)
# ==========================================================

_RX_TE = re.compile(
    r"(?i)(?<![A-Z0-9])TE\s*[-:\s–—]*\s*(\d{1,2})\s*/\s*(\d{1,2})(?:\s*/\s*(\d{2,4}))?"
)
_RX_EJA = re.compile(
    r"(?i)(?<![A-Z0-9])(TE|MC|MCC)\s*[-:\s–—]*\s*(\d{1,2})\s*/\s*(\d{1,2})(?:\s*/\s*(\d{2,4}))?"
)

def _is_missing_text(v) -> bool:
    s = _safe_str(v)
    if not s:
        return True
    return s.lower() in {"0", "-", "nan", "none", "null"}

def _parse_user_date(date_str: str) -> Optional[datetime]:
    s = _safe_str(date_str)
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None

def _parse_period_date(date_str: str, label: str) -> datetime:
    """
    Mantém compatibilidade com input type=date (YYYY-MM-DD) e aceita dd/mm/aa|aaaa.
    """
    s = _safe_str(date_str)
    if not s:
        raise ValueError(f"Informe {label}.")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    raise ValueError(f"Formato de {label} inválido: '{date_str}'.")

def _extract_te_date_from_text(text: str, period_start: datetime, period_end: datetime):
    """
    Extrai TE - dd/mm[/aa|aaaa] de um texto (OBS).
    Ano ausente: tenta encaixar no período; se não der, assume ano corrente.
    Retorna (dt, match_txt, year_inferred).
    """
    s = _safe_str(text)
    if not s:
        return None, None, False

    m = _RX_TE.search(s)
    if not m:
        return None, None, False

    day = int(m.group(1))
    month = int(m.group(2))
    year_raw = m.group(3)

    year_inferred = False
    if year_raw:
        y = int(year_raw)
        if y < 100:
            y += 2000
        years_to_try = [y]
    else:
        year_inferred = True
        years_to_try = [period_start.year]
        if period_end.year != period_start.year:
            years_to_try.append(period_end.year)

    for y in years_to_try:
        try:
            dt = datetime(y, month, day)
        except Exception:
            continue
        if period_start <= dt <= period_end:
            return dt, m.group(0), year_inferred

    if not year_raw:
        y = datetime.now().year
        try:
            return datetime(y, month, day), m.group(0), year_inferred
        except Exception:
            return None, m.group(0), year_inferred

    for y in years_to_try:
        try:
            return datetime(y, month, day), m.group(0), year_inferred
        except Exception:
            continue

    return None, m.group(0), year_inferred

def _label_set(ws, addr: str, label: str, value: str):
    """
    Mantém prefixo do template quando existir (ex.: 'Unidade Escolar: ...').
    """
    current = ws[addr].value
    val = _safe_str(value)

    if isinstance(current, str) and ":" in current:
        left = current.split(":", 1)[0].strip()
        if _norm_header_compact(left) == _norm_header_compact(label):
            set_merged_cell_value(ws, addr, f"{left}: {val}")
            return

    set_merged_cell_value(ws, addr, val)


@app.route("/quadros/transferencias", methods=["GET", "POST"])
@login_required
def quadro_transferencias():
    if request.method == "POST":
        enable_eja = _is_eja_enabled()

        period_start_str = request.form.get("period_start")
        period_end_str = request.form.get("period_end")

        responsavel = request.form.get("responsavel")
        diretor_nome = request.form.get("diretor_nome") or "Luciana Rocha Augustinho"
        data_quadro_in = request.form.get("data_quadro")  # opcional

        fundamental_file = request.files.get("lista_fundamental")
        eja_file = request.files.get("lista_eja")

        if not period_start_str or not period_end_str or not responsavel:
            flash("Por favor, preencha todos os campos.", "error")
            return redirect(url_for("quadro_transferencias"))

        # FUNDAMENTAL
        if fundamental_file and fundamental_file.filename != "":
            fundamental_filename = secure_filename(
                f"fundamental_{uuid.uuid4().hex}_" + fundamental_file.filename
            )
            upload_folder = current_app.config.get("UPLOAD_FOLDER", "uploads")
            fundamental_path = os.path.join(upload_folder, fundamental_filename)
            fundamental_file.save(fundamental_path)
            session["lista_fundamental"] = fundamental_path
        else:
            fundamental_path = session.get("lista_fundamental")
            if not fundamental_path or not os.path.exists(fundamental_path):
                flash("Lista Piloto Fundamental não encontrada.", "error")
                return redirect(url_for("quadro_transferencias"))

        # EJA opcional
        eja_path = None
        if enable_eja and eja_file and eja_file.filename != "":
            eja_filename = secure_filename(f"eja_{uuid.uuid4().hex}_" + eja_file.filename)
            upload_folder = current_app.config.get("UPLOAD_FOLDER", "uploads")
            eja_path = os.path.join(upload_folder, eja_filename)
            eja_file.save(eja_path)
            session["lista_eja"] = eja_path
        elif enable_eja:
            eja_path = session.get("lista_eja")

        try:
            period_start = _parse_period_date(period_start_str, "a data inicial")
            period_end = _parse_period_date(period_end_str, "a data final")
        except ValueError as e:
            flash(str(e), "error")
            return redirect(url_for("quadro_transferencias"))

        if period_end < period_start:
            flash("A data final não pode ser menor que a data inicial.", "error")
            return redirect(url_for("quadro_transferencias"))

        data_quadro_dt = _parse_user_date(data_quadro_in) or datetime.now()

        # Lê LISTA CORRIDA (Fundamental)
        try:
            df_fundamental = pd.read_excel(fundamental_path, sheet_name="LISTA CORRIDA")
        except Exception as e:
            flash(f"Erro ao ler a Lista Piloto Fundamental: {str(e)}", "error")
            return redirect(url_for("quadro_transferencias"))

        colmap = _build_colmap(df_fundamental)

        col_serie = _pick_col(colmap, "SÉRIE", "SERIE")
        col_nome = _pick_col(colmap, "NOME")
        col_dn = _pick_col(colmap, "DATA NASC.", "DATA NASC", "DATANASC")
        col_ra = _pick_col(colmap, "RA")
        col_obs = _pick_col(colmap, "OBS", "OBSERVACAO", "OBSERVAÇÃO")
        col_local_te = _pick_col(colmap, "LOCAL TE", "LOCALTE")

        if not col_nome or not col_dn or not col_ra or not col_serie or not col_obs:
            flash("A aba 'LISTA CORRIDA' não contém cabeçalhos essenciais (SÉRIE, NOME, DATA NASC., RA, OBS).", "error")
            return redirect(url_for("quadro_transferencias"))

        debug = []
        debug.append("[quadro_transferencias] Aba lida: LISTA CORRIDA (Fundamental)")
        debug.append(
            f"[quadro_transferencias] Colunas detectadas: "
            f"SÉRIE='{col_serie}', NOME='{col_nome}', DATA NASC.='{col_dn}', RA='{col_ra}', "
            f"OBS='{col_obs}', LOCAL TE='{col_local_te}'"
        )

        transfer_records = []
        invalid_te_dates = 0

        use_cols = [col_serie, col_nome, col_dn, col_ra, col_obs]
        if col_local_te and col_local_te not in use_cols:
            use_cols.append(col_local_te)

        df_sub = df_fundamental[use_cols].copy()

        for row in df_sub.itertuples(index=False, name=None):
            row_dict = dict(zip(df_sub.columns, row))

            te_dt, te_match_txt, _ = _extract_te_date_from_text(
                row_dict.get(col_obs), period_start, period_end
            )

            if te_match_txt and not te_dt:
                invalid_te_dates += 1
                continue

            if not te_dt:
                continue

            if not (period_start <= te_dt <= period_end):
                continue

            nome = _safe_str(row_dict.get(col_nome))

            dn_val = row_dict.get(col_dn)
            dn_str = ""
            if pd.notna(dn_val):
                try:
                    dn_dt = pd.to_datetime(dn_val, errors="coerce")
                    dn_str = dn_dt.strftime("%d/%m/%Y") if pd.notna(dn_dt) else ""
                except Exception:
                    dn_str = ""

            ra = _safe_str(row_dict.get(col_ra))
            nivel_classe = _safe_str(row_dict.get(col_serie))

            # FIX H/I/J:
            local_te_raw = _safe_str(row_dict.get(col_local_te)) if col_local_te else ""
            local_te = "-" if _is_missing_text(local_te_raw) else local_te_raw

            record = {
                "nome": nome,
                "dn": dn_str,
                "ra": ra,
                "situacao": "Parcial",
                "breda": "Não",
                "nivel_classe": nivel_classe,
                "tipo": "TE",
                "observacao": local_te,              # H
                "remanejamento": "-",                # I
                "data": te_dt.strftime("%d/%m/%Y"),  # J
            }
            transfer_records.append(record)

        debug.append(f"[quadro_transferencias] TE válidos no período: {len(transfer_records)}")
        debug.append(f"[quadro_transferencias] Datas TE inválidas descartadas: {invalid_te_dates}")

        # EJA opcional (mantido)
        if enable_eja and eja_path and os.path.exists(eja_path):
            try:
                df_eja = pd.read_excel(eja_path, sheet_name="LISTA CORRIDA")
            except Exception as e:
                flash(f"Erro ao ler a Lista Piloto EJA: {str(e)}", "error")
                return redirect(url_for("quadro_transferencias"))

            colmap_eja = _build_colmap(df_eja)
            eja_col_nome = _pick_col(colmap_eja, "NOME")
            eja_col_dn = _pick_col(colmap_eja, "DATA NASC.", "DATA NASC")
            eja_col_ra = _pick_col(colmap_eja, "RA")
            eja_col_serie = _pick_col(colmap_eja, "SÉRIE", "SERIE")
            eja_col_obs = _pick_col(colmap_eja, "OBS")
            eja_col_local_te = _pick_col(colmap_eja, "LOCAL TE", "LOCALTE")

            if eja_col_nome and eja_col_ra and eja_col_serie and eja_col_obs:
                df_eja_sub = df_eja[
                    [c for c in [eja_col_serie, eja_col_nome, eja_col_dn, eja_col_ra, eja_col_obs, eja_col_local_te] if c]
                ].copy()

                for row in df_eja_sub.itertuples(index=False, name=None):
                    row_dict = dict(zip(df_eja_sub.columns, row))
                    txt = _safe_str(row_dict.get(eja_col_obs))

                    m = _RX_EJA.search(txt)
                    if not m:
                        continue

                    tipo_str = m.group(1).upper()
                    day = int(m.group(2))
                    month = int(m.group(3))
                    year_raw = m.group(4)

                    if year_raw:
                        y = int(year_raw)
                        if y < 100:
                            y += 2000
                    else:
                        y = period_start.year

                    try:
                        dt = datetime(y, month, day)
                    except Exception:
                        continue

                    if not (period_start <= dt <= period_end):
                        continue

                    nome = _safe_str(row_dict.get(eja_col_nome))
                    ra = _safe_str(row_dict.get(eja_col_ra))
                    nivel_classe = _safe_str(row_dict.get(eja_col_serie))

                    dn_str = ""
                    if eja_col_dn:
                        dn_val = row_dict.get(eja_col_dn)
                        if pd.notna(dn_val):
                            try:
                                dn_dt = pd.to_datetime(dn_val, errors="coerce")
                                dn_str = dn_dt.strftime("%d/%m/%Y") if pd.notna(dn_dt) else ""
                            except Exception:
                                dn_str = ""

                    local_te_raw = _safe_str(row_dict.get(eja_col_local_te)) if eja_col_local_te else ""
                    local_te = "-" if _is_missing_text(local_te_raw) else local_te_raw

                    transfer_records.append({
                        "nome": nome,
                        "dn": dn_str,
                        "ra": ra,
                        "situacao": "Parcial",
                        "breda": "Não",
                        "nivel_classe": nivel_classe,
                        "tipo": tipo_str,
                        "observacao": local_te,           # H
                        "remanejamento": "-",             # I
                        "data": dt.strftime("%d/%m/%Y"),  # J
                    })

        if not transfer_records:
            flash("Nenhum registro de TE/MC/MCC encontrado no período especificado.", "error")
            try:
                current_app.logger.info("\n".join(debug))
            except Exception:
                print("\n".join(debug))
            return redirect(url_for("quadro_transferencias"))

        model_path = os.path.join("modelos", "Quadro Informativo - Modelo.xlsx")
        if not os.path.exists(model_path):
            flash("Modelo de Quadro Informativo (Transferências) não encontrado.", "error")
            return redirect(url_for("quadro_transferencias"))

        try:
            with open(model_path, "rb") as f:
                wb = load_workbook(f, data_only=False)
        except Exception as e:
            flash(f"Erro ao ler o modelo: {str(e)}", "error")
            return redirect(url_for("quadro_transferencias"))

        ws = wb.active

        SCHOOL_NAME = "E.M José Padin Mouta"
        _label_set(ws, "A7", "Unidade Escolar", SCHOOL_NAME)
        _label_set(ws, "A8", "Diretor(a)", diretor_nome)

        set_merged_cell_value(ws, "B9", responsavel)
        set_merged_cell_value(ws, "J9", data_quadro_dt.strftime("%d/%m/%Y"))

        current_row = 12
        for record in transfer_records:
            set_merged_cell_value(ws, f"A{current_row}", record["nome"])
            set_merged_cell_value(ws, f"B{current_row}", record["dn"])
            set_merged_cell_value(ws, f"C{current_row}", record["ra"])
            set_merged_cell_value(ws, f"D{current_row}", record["situacao"])
            set_merged_cell_value(ws, f"E{current_row}", record["breda"])
            set_merged_cell_value(ws, f"F{current_row}", record["nivel_classe"])
            set_merged_cell_value(ws, f"G{current_row}", record["tipo"])

            # FIX H/I/J
            set_merged_cell_value(ws, f"H{current_row}", record["observacao"])
            set_merged_cell_value(ws, f"I{current_row}", "-")
            set_merged_cell_value(ws, f"J{current_row}", record["data"])

            current_row += 1

        debug.append(f"[quadro_transferencias] Linhas preenchidas no modelo: {len(transfer_records)} (início A12)")
        try:
            current_app.logger.info("\n".join(debug))
        except Exception:
            print("\n".join(debug))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Quadro_de_Transferencias_{period_start.strftime('%d%m')}_{period_end.strftime('%d%m')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    return render_template("quadro_transferencias.html")


# ==========================================================
#  QUADRO – QUANTITATIVO MENSAL (Fundamental)
#  (mantido: parse flexível do período + debug sheet oculta)
# ==========================================================

_RX_ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_RX_BR_DATE = re.compile(r"^\s*(\d{1,2})\s*/\s*(\d{1,2})(?:\s*/\s*(\d{2,4}))?\s*$")

def parse_date_flexible(value: str, *, default_year: Optional[int] = None, field_label: str = "data") -> datetime:
    """
    Aceita:
      - YYYY-MM-DD (input type=date)
      - dd/mm/aaaa
      - dd/mm/aa
      - dd/mm (assume ano corrente, ou default_year se informado)
    """
    if value is None or str(value).strip() == "":
        raise ValueError(f"Informe {field_label}.")

    s = str(value).strip()

    if _RX_ISO_DATE.match(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except Exception:
            raise ValueError(f"{field_label.capitalize()} inválida: '{value}'.")

    m = _RX_BR_DATE.match(s)
    if not m:
        raise ValueError(
            f"{field_label.capitalize()} inválida: '{value}'. Use 16/01, 16/01/26, 16/01/2026 (ou selecione no calendário)."
        )

    day = int(m.group(1))
    month = int(m.group(2))
    y_str = m.group(3)

    if not y_str:
        year = int(default_year) if default_year is not None else datetime.now().year
    else:
        year = int(y_str)
        if year < 100:
            year += 2000

    try:
        return datetime(year, month, day)
    except ValueError:
        raise ValueError(f"{field_label.capitalize()} inválida: '{value}' (dia/mês não existe).")


_RX_TE_DATE_FLEX = re.compile(
    r"\bTE\b\s*[-:–—]?\s*(\d{1,2})\s*/\s*(\d{1,2})(?:\s*/\s*(\d{2,4}))?\b",
    re.IGNORECASE,
)

def detect_te_date_from_obs_flexible(
    obs_text,
    *,
    default_year: Optional[int] = None,
) -> Tuple[Optional[datetime], Optional[str], Optional[str], bool]:
    """
    Procura TE + data em OBS.
    Aceita: TE - 16/01, TE - 16/01/26, TE - 16/01/2026.
    Se faltar ano, assume ano corrente (ou default_year se fornecido).
    Retorna: (dt, regra, trecho_match, year_inferred)
    """
    if obs_text is None:
        return None, None, None, False

    text = str(obs_text).strip()
    if text == "":
        return None, None, None, False

    m = _RX_TE_DATE_FLEX.search(text)
    if not m:
        return None, None, None, False

    day = int(m.group(1))
    month = int(m.group(2))
    y_str = m.group(3)

    year_inferred = False
    if not y_str:
        year = int(default_year) if default_year is not None else datetime.now().year
        year_inferred = True
    else:
        year = int(y_str)
        if year < 100:
            year += 2000

    try:
        dt = datetime(year, month, day)
    except ValueError:
        return None, "OBS:TE_DATE", m.group(0), year_inferred

    return dt, "OBS:TE_DATE", m.group(0), year_inferred


def _serie_key_from_value(serie_val: str) -> Optional[str]:
    """Extrai 2º/3º/4º/5º de valores como '4ºF', '5º D', etc."""
    s = "" if serie_val is None else str(serie_val).strip()
    m = re.search(r"(\d)\s*º", s)
    if not m:
        return None
    n = m.group(1)
    if n in {"2", "3", "4", "5"}:
        return f"{n}º"
    return None


def _normalize_tipo_te(val) -> str:
    """
    Normaliza o TIPO TE para bater no mapping do modelo.
    Fora disso vira 'Sem Informação' (sem inventar dado).
    """
    if _is_missing_value(val):
        return "Sem Informação"

    raw = str(val).strip()
    norm = _norm_header_compact(raw)

    if "DENTRO" in norm or "REDEMUNICIPAL" in norm or "MUNICIPAL" in norm:
        return "Dentro da Rede"
    if "ESTAD" in norm:
        return "Rede Estadual"
    if "LITORAL" in norm or "BAIXADA" in norm:
        return "Litoral"
    if "MUDANCA" in norm and "MUNICIP" in norm:
        return "Mudança de Municipio"
    if "SAOPAULO" in norm:
        return "São Paulo"
    if "ABCD" in norm:
        return "ABCD"
    if "INTERIOR" in norm:
        return "Interior"
    if "OUTROSESTADOS" in norm or ("OUTROS" in norm and "ESTAD" in norm):
        return "Outros Estados"
    if "PARTICULAR" in norm:
        return "Particular"
    if "PAIS" in norm:
        return "País"
    if "SEMINFORMA" in norm:
        return "Sem Informação"

    return raw


def _recreate_debug_sheet_hidden(wb, title: str = "DEBUG_TE"):
    """Cria/zera uma aba de debug (oculta) no workbook."""
    if title in wb.sheetnames:
        wb.remove(wb[title])
    ws_dbg = wb.create_sheet(title)
    ws_dbg.sheet_state = "hidden"
    ws_dbg.append(
        [
            "LINHA_ARQUIVO",
            "RM",
            "NOME",
            "SERIE",
            "OBS_ORIGINAL",
            "TE_DATA_EXTRAIDA",
            "ANO_INFERIDO",
            "TRECHO_MATCH",
            "STATUS",
            "MOTIVO",
            "TIPO_TE_RAW",
            "TIPO_TE_NORMALIZADO",
        ]
    )
    return ws_dbg


@app.route("/quadros/quantitativo_mensal", methods=["GET", "POST"])
@login_required
def quadro_quantitativo_mensal():
    if request.method == "POST":
        period_start_str = request.form.get("period_start")
        period_end_str = request.form.get("period_end")
        responsavel = request.form.get("responsavel")
        mes_ano = request.form.get("mes_ano")

        if not responsavel or not str(responsavel).strip():
            flash("Preencha o campo Responsável.", "error")
            return redirect(url_for("quadro_quantitativo_mensal"))

        default_year = datetime.now().year

        try:
            period_start = parse_date_flexible(period_start_str, default_year=default_year, field_label="a data inicial")
            period_end = parse_date_flexible(period_end_str, default_year=default_year, field_label="a data final")
        except ValueError as e:
            flash(str(e), "error")
            return redirect(url_for("quadro_quantitativo_mensal"))

        if period_end < period_start:
            flash("A data final não pode ser menor que a data inicial.", "error")
            return redirect(url_for("quadro_quantitativo_mensal"))

        if not mes_ano or not str(mes_ano).strip():
            meses = {
                1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
                5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
                9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
            }
            mes_ano = f"{meses[datetime.now().month]}/{datetime.now().year}"
        mes_ano = str(mes_ano).strip()

        fundamental_file = request.files.get("lista_fundamental")
        if fundamental_file and fundamental_file.filename:
            fundamental_filename = secure_filename(
                f"fundamental_{uuid.uuid4().hex}_" + fundamental_file.filename
            )
            upload_folder = current_app.config.get("UPLOAD_FOLDER", "uploads")
            fundamental_path = os.path.join(upload_folder, fundamental_filename)
            fundamental_file.save(fundamental_path)
            session["lista_fundamental"] = fundamental_path
        else:
            fundamental_path = session.get("lista_fundamental")
            if not fundamental_path or not os.path.exists(fundamental_path):
                flash("Arquivo da Lista Piloto Fundamental não encontrado.", "error")
                return redirect(url_for("quadro_quantitativo_mensal"))

        try:
            df = pd.read_excel(fundamental_path, sheet_name="LISTA CORRIDA")
        except Exception as e:
            flash(f"Erro ao ler a Lista Piloto Fundamental: {str(e)}", "error")
            return redirect(url_for("quadro_quantitativo_mensal"))

        col_rm = _find_df_col(df, ["RM"])
        col_nome = _find_df_col(df, ["NOME"])
        col_serie = _find_df_col(df, ["SÉRIE", "SERIE"])
        col_obs = _find_df_col(df, ["OBS"])
        col_tipo_te = _find_df_col(df, ["TIPO TE", "TIPO_TE", "TIPO  TE"])

        if not col_serie or not col_obs:
            flash("Não foi possível localizar as colunas essenciais (SÉRIE e/ou OBS) na LISTA CORRIDA.", "error")
            return redirect(url_for("quadro_quantitativo_mensal"))

        model_path = os.path.join("modelos", "Quadro Quantitativo Mensal - Modelo.xlsx")
        if not os.path.exists(model_path):
            flash("Modelo de Quadro Quantitativo Mensal não encontrado.", "error")
            return redirect(url_for("quadro_quantitativo_mensal"))

        try:
            with open(model_path, "rb") as f:
                wb = load_workbook(f, data_only=False)
        except Exception as e:
            flash(f"Erro ao ler o modelo: {str(e)}", "error")
            return redirect(url_for("quadro_quantitativo_mensal"))

        ws = wb.active

        mapping = {
            "2º": {
                "Dentro da Rede": "K14",
                "Rede Estadual": "K15",
                "Litoral": "K16",
                "Mudança de Municipio": "K16",
                "São Paulo": "K17",
                "ABCD": "K18",
                "Interior": "K19",
                "Outros Estados": "K20",
                "Particular": "K21",
                "País": "K22",
                "Sem Informação": "K23",
            },
            "3º": {
                "Dentro da Rede": "L14",
                "Rede Estadual": "L15",
                "Litoral": "L16",
                "Mudança de Municipio": "L16",
                "São Paulo": "L17",
                "ABCD": "L18",
                "Interior": "L19",
                "Outros Estados": "L20",
                "Particular": "L21",
                "País": "L22",
                "Sem Informação": "L23",
            },
            "4º": {
                "Dentro da Rede": "M14",
                "Rede Estadual": "M15",
                "Litoral": "M16",
                "Mudança de Municipio": "M16",
                "São Paulo": "M17",
                "ABCD": "M18",
                "Interior": "M19",
                "Outros Estados": "M20",
                "Particular": "M21",
                "País": "M22",
                "Sem Informação": "M23",
            },
            "5º": {
                "Dentro da Rede": "N14",
                "Rede Estadual": "N15",
                "Litoral": "N16",
                "Mudança de Municipio": "N16",
                "São Paulo": "N17",
                "ABCD": "N18",
                "Interior": "N19",
                "Outros Estados": "N20",
                "Particular": "N21",
                "País": "N22",
                "Sem Informação": "N23",
            },
        }

        # Zera determinísticamente todas as células-alvo
        for tipos in mapping.values():
            for cell_addr in tipos.values():
                set_merged_cell_value(ws, cell_addr, 0)

        ws_dbg = _recreate_debug_sheet_hidden(wb, "DEBUG_TE")

        counted = 0
        discarded = 0

        for i, row in df.iterrows():
            linha_arquivo = int(i) + 2

            rm = row.get(col_rm) if col_rm else None
            nome = row.get(col_nome) if col_nome else None
            serie_val = row.get(col_serie, "")
            obs_val = row.get(col_obs, "")

            if _is_missing_value(obs_val):
                continue

            te_dt, rule, match_txt, year_inferred = detect_te_date_from_obs_flexible(
                obs_val,
                default_year=default_year,
            )

            if not match_txt:
                continue

            if not te_dt:
                discarded += 1
                ws_dbg.append(
                    [
                        linha_arquivo,
                        "" if rm is None else str(rm),
                        "" if nome is None else str(nome),
                        "" if serie_val is None else str(serie_val),
                        str(obs_val).strip(),
                        "",
                        "SIM" if year_inferred else "NAO",
                        match_txt,
                        "SKIPPED",
                        "Data TE inválida em OBS",
                        "" if col_tipo_te is None else _safe_str(row.get(col_tipo_te)),
                        "" if col_tipo_te is None else _normalize_tipo_te(row.get(col_tipo_te)),
                    ]
                )
                continue

            if not (period_start <= te_dt <= period_end):
                discarded += 1
                ws_dbg.append(
                    [
                        linha_arquivo,
                        "" if rm is None else str(rm),
                        "" if nome is None else str(nome),
                        "" if serie_val is None else str(serie_val),
                        str(obs_val).strip(),
                        te_dt.strftime("%d/%m/%Y"),
                        "SIM" if year_inferred else "NAO",
                        match_txt,
                        "SKIPPED",
                        "Fora do período informado",
                        "" if col_tipo_te is None else _safe_str(row.get(col_tipo_te)),
                        "" if col_tipo_te is None else _normalize_tipo_te(row.get(col_tipo_te)),
                    ]
                )
                continue

            serie_key = _serie_key_from_value(serie_val)
            if not serie_key or serie_key not in mapping:
                discarded += 1
                ws_dbg.append(
                    [
                        linha_arquivo,
                        "" if rm is None else str(rm),
                        "" if nome is None else str(nome),
                        "" if serie_val is None else str(serie_val),
                        str(obs_val).strip(),
                        te_dt.strftime("%d/%m/%Y"),
                        "SIM" if year_inferred else "NAO",
                        match_txt,
                        "SKIPPED",
                        "Série fora de 2º–5º ou ilegível",
                        "" if col_tipo_te is None else _safe_str(row.get(col_tipo_te)),
                        "" if col_tipo_te is None else _normalize_tipo_te(row.get(col_tipo_te)),
                    ]
                )
                continue

            tipo_raw = row.get(col_tipo_te, None) if col_tipo_te else None
            tipo_te = _normalize_tipo_te(tipo_raw)
            if tipo_te not in mapping[serie_key]:
                tipo_te = "Sem Informação"

            cell_addr = mapping[serie_key][tipo_te]
            current_val = ws[cell_addr].value
            current_val = current_val if isinstance(current_val, (int, float)) else 0
            set_merged_cell_value(ws, cell_addr, current_val + 1)

            counted += 1
            ws_dbg.append(
                [
                    linha_arquivo,
                    "" if rm is None else str(rm),
                    "" if nome is None else str(nome),
                    "" if serie_val is None else str(serie_val),
                    str(obs_val).strip(),
                    te_dt.strftime("%d/%m/%Y"),
                    "SIM" if year_inferred else "NAO",
                    match_txt,
                    "COUNTED",
                    "",
                    "" if tipo_raw is None else str(tipo_raw),
                    tipo_te,
                ]
            )

        set_merged_cell_value(ws, "B3", str(responsavel).strip())
        set_merged_cell_value(ws, "D3", f"{period_start.strftime('%d/%m/%Y')} a {period_end.strftime('%d/%m/%Y')}")

        with _temp_unprotect_sheet(ws):
            set_merged_cell_value(ws, "A6", "E.M José Padin Mouta")
            set_merged_cell_value(ws, "A8", mes_ano)
            set_merged_cell_value(ws, "A10", "QUADRO GERAL DE TRANSFERENCIAS EXPEDIDAS - 2026")

        current_app.logger.info(
            "[QUADRO_QUANTITATIVO] periodo=%s..%s | counted=%s | discarded=%s | ano_padrao_sem_ano=%s",
            period_start.strftime("%d/%m/%Y"),
            period_end.strftime("%d/%m/%Y"),
            counted,
            discarded,
            default_year,
        )

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Quadro_Quantitativo_Fundamental_{period_start.strftime('%d%m')}_{period_end.strftime('%d%m')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    meses = {
        1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
        5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
        9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
    }
    default_mes_ano = f"{meses[datetime.now().month]}/{datetime.now().year}"
    return render_template("quadro_quantitativo_mensal.html", default_mes_ano=default_mes_ano)

# ==========================================================
#  QUADRO QUANTITATIVO DE INCLUSÃO – REGULAR (EJA DESCONSIDERADA)
#  - Contagem por turma via LISTA CORRIDA
#  - Regras:
#      * Turma: Coluna A
#      * RM (identificador único): Coluna C (NÃO contabiliza vazio/0)
#      * Inclusão: Coluna N == "Sim" (case/trim)
#      * Plano de Ação: Coluna P com profissional válido (ignora vazio/0/-)
#      * Profissionais: únicos por turma (dedupe case-insensitive e colapsando espaços)
#  - Alerta: turmas com 2+ profissionais distintos (Plano de Ação)
#  - Preenchimento por mapeamento automático do MODELO
# ==========================================================

import os
import re
import json
import base64
import uuid
from io import BytesIO
from collections import defaultdict
from datetime import datetime

from flask import (
    request,
    flash,
    redirect,
    url_for,
    send_file,
    render_template,
)
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

# OBS: este código assume que você já tem:
# - app = Flask(__name__)
# - app.config["UPLOAD_FOLDER"]
# no seu arquivo principal.


def _collapse_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def _normalize_rm(v) -> str:
    """
    Normaliza RM para string comparável e SEGURA para deduplicação.
    Regras importantes:
      - RM vazio/None => ""
      - RM 0 / "0" / "0000" / "0.0" => "" (NÃO conta)
      - Remove caracteres não numéricos e remove zeros à esquerda
        (ex.: "00012" -> "12")
      - Float não-inteiro (ex.: 12.5) => "" (inválido)
    """
    if v is None:
        return ""

    # int direto
    if isinstance(v, int):
        if v == 0:
            return ""
        return str(v).strip()

    # float comum no Excel
    if isinstance(v, float):
        if not v.is_integer():
            return ""  # RM com decimal é inválido
        iv = int(v)
        if iv == 0:
            return ""
        return str(iv).strip()

    s = _collapse_spaces(str(v))
    if not s:
        return ""

    # casos de "nan"/"none" etc
    if s.casefold() in {"nan", "none"}:
        return ""

    # extrai somente dígitos (tolerante a RM com espaços/pontos)
    digits = re.sub(r"\D+", "", s)
    if not digits:
        return ""

    # remove zeros à esquerda
    digits = digits.lstrip("0")
    if not digits:
        return ""  # era tudo zero

    return digits


def _normalize_turma(v) -> str:
    """
    Normaliza turma para padrão: '2ºA', '3ºC' etc.
    Aceita entradas comuns: '2ºA', '2-A', '2 A', '2a', '2ªA' etc.
    """
    if v is None:
        return ""
    s = _collapse_spaces(str(v))
    if not s:
        return ""

    s = s.replace("ª", "").replace("º", "").replace("°", "")
    s = s.replace("-", "").replace("/", "").replace("\\", "")
    s = s.replace(" ", "")
    s = s.upper()

    m = re.match(r"^(\d{1,2})([A-Z])$", s)
    if not m:
        return ""
    num = int(m.group(1))
    letra = m.group(2)
    return f"{num}º{letra}"


def _is_sim(v) -> bool:
    """Coluna N: deve ser exatamente 'Sim' (tolerante a caixa/trim)."""
    s = _collapse_spaces("" if v is None else str(v))
    return s.casefold() == "sim"


def _is_valid_prof(v) -> bool:
    """Coluna P: profissional válido (ignora vazio, 0, -)."""
    if v is None:
        return False
    s = _collapse_spaces(str(v))
    if not s:
        return False
    if s in {"0", "-"}:
        return False
    if s.casefold() in {"nan", "none"}:
        return False
    return True


def _prof_key(v: str) -> str:
    """Chave de deduplicação: strip, colapsa espaços e compara case-insensitive."""
    return _collapse_spaces(v).casefold()


def _build_template_map(ws_model):
    """
    Mapeia automaticamente o MODELO:
      - Encontra turmas nos rótulos (colunas B, F, J) como '2-A', '3-C', '5-F'...
      - Para cada turma, calcula as células de quantidade:
          Inclusão:        D/H/L na linha do rótulo
          Plano de Ação:   D/H/L na linha do rótulo+1
          Profissionais:   D/H/L na linha do rótulo+2
    Retorna:
      dict { '2ºA': {'inc_qtd': 'D13', 'plano_qtd': 'D14', 'prof_qtd': 'D15'}, ... }
    """
    label_to_qty = {2: 4, 6: 8, 10: 12}  # B->D, F->H, J->L
    turma_cells = {}

    max_row = ws_model.max_row
    for label_col, qty_col in label_to_qty.items():
        for r in range(1, max_row + 1):
            v = ws_model.cell(r, label_col).value
            if not isinstance(v, str):
                continue
            raw = _collapse_spaces(v)
            if not re.match(r"^\d{1,2}\s*-\s*[A-Za-z]$", raw):
                continue

            turma_norm = _normalize_turma(raw)
            if not turma_norm:
                continue

            inc_row = r
            plano_row = r + 1
            prof_row = r + 2

            turma_cells[turma_norm] = {
                "inc_qtd": ws_model.cell(inc_row, qty_col).coordinate,
                "plano_qtd": ws_model.cell(plano_row, qty_col).coordinate,
                "prof_qtd": ws_model.cell(prof_row, qty_col).coordinate,
            }

    return turma_cells


def _collect_counts_from_lista_corrida(ws_lista, valid_turmas):
    """
    Lê LISTA CORRIDA e retorna:
      - inc_counts[turma] = qtd alunos inclusão (N == 'Sim'), dedupe por RM
      - plano_counts[turma] = qtd alunos com plano (P válido), dedupe por RM
      - profs_by_turma[turma] = dict key-> {'display': str, 'alunos': [(rm, nome), ...]}
    Colunas (0-index):
      A turma = 0
      C RM    = 2
      D nome  = 3 (opcional p/ auditoria)
      N incl  = 13
      P prof  = 15
    Observação crítica:
      - RM vazio/0 NÃO é contabilizado (alunos não matriculados).
    """
    inc_rms = defaultdict(set)
    plano_rms = defaultdict(set)
    profs_by_turma = defaultdict(lambda: defaultdict(lambda: {"display": "", "alunos": []}))

    for row in ws_lista.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        turma = _normalize_turma(row[0] if len(row) > 0 else None)
        if not turma or (valid_turmas and turma not in valid_turmas):
            continue

        rm = _normalize_rm(row[2] if len(row) > 2 else None)

        # >>> CORREÇÃO PRINCIPAL: ignora RM inválido (vazio/0/etc)
        if not rm:
            continue

        nome = _collapse_spaces(str(row[3])) if len(row) > 3 and row[3] is not None else ""

        # Inclusão: coluna N (index 13) == "Sim"
        incl_val = row[13] if len(row) > 13 else None
        if _is_sim(incl_val):
            inc_rms[turma].add(rm)

        # Plano de Ação / Profissionais: coluna P (index 15)
        prof_val = row[15] if len(row) > 15 else None
        if _is_valid_prof(prof_val):
            plano_rms[turma].add(rm)

            display = _collapse_spaces(str(prof_val))
            key = _prof_key(display)
            bucket = profs_by_turma[turma][key]

            if not bucket["display"]:
                bucket["display"] = display

            # auditoria leve
            bucket["alunos"].append((rm, nome))

    inc_counts = {t: len(rms) for t, rms in inc_rms.items()}
    plano_counts = {t: len(rms) for t, rms in plano_rms.items()}

    return inc_counts, plano_counts, profs_by_turma


@app.route("/quantinclusao", methods=["GET", "POST"])
def quantinclusao():
    if request.method == "POST":
        # EJA foi desconsiderada conforme solicitado.
        reg_file = (
            request.files.get("lista_regular")
            or request.files.get("lista_fundamental")
            or request.files.get("lista")
        )
        responsavel = request.form.get("responsavel")

        if not reg_file or reg_file.filename == "":
            flash("Selecione o arquivo da Lista Piloto (Regular/Fundamental).", "error")
            return redirect(url_for("quantinclusao"))

        if not responsavel or responsavel.strip() == "":
            flash("Informe o Responsável pelo preenchimento.", "error")
            return redirect(url_for("quantinclusao"))

        # Salva upload
        reg_filename = secure_filename(f"regular_{uuid.uuid4().hex}_{reg_file.filename}")
        reg_path = os.path.join(app.config["UPLOAD_FOLDER"], reg_filename)
        reg_file.save(reg_path)

        # Abre LISTA CORRIDA (read_only p/ performance)
        try:
            wb_reg = load_workbook(reg_path, data_only=True, read_only=True)
            ws_lista_reg = wb_reg["LISTA CORRIDA"]
        except Exception as e:
            flash(f"Erro ao ler o arquivo: {str(e)}", "error")
            return redirect(url_for("quantinclusao"))

        # Abre modelo e monta mapeamento automático
        model_path = os.path.join("modelos", "Quadro Quantitativo de Inclusão - Modelo.xlsx")
        try:
            wb_model = load_workbook(model_path, data_only=False)
            ws_model = wb_model.active
        except Exception as e:
            flash(f"Erro ao abrir o modelo de inclusão: {str(e)}", "error")
            return redirect(url_for("quantinclusao"))

        template_map = _build_template_map(ws_model)
        valid_turmas = set(template_map.keys())

        # Contagens por turma (somente turmas existentes no modelo)
        inc_counts, plano_counts, profs_by_turma = _collect_counts_from_lista_corrida(
            ws_lista_reg, valid_turmas
        )

        # Preenche as 3 linhas por turma: Inclusão / Plano / Profissionais
        for turma, cells in template_map.items():
            inc = inc_counts.get(turma, 0)
            plano = plano_counts.get(turma, 0)
            profs = len(profs_by_turma.get(turma, {}))

            ws_model[cells["inc_qtd"]] = inc
            ws_model[cells["plano_qtd"]] = plano
            ws_model[cells["prof_qtd"]] = profs

        # Cabeçalho
        meses = {
            1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL",
            5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO",
            9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO",
        }
        now = datetime.now()
        mes_ano = f"{meses[now.month]}/{now.year}"

        # B4: troca apenas quando reconhece padrão; senão, anexa com cuidado
        try:
            b4 = ws_model["B4"].value or ""
            b4s = str(b4)
            if re.search(r"MÊS\s*/\s*\d{4}", b4s, flags=re.IGNORECASE):
                ws_model["B4"] = re.sub(r"MÊS\s*/\s*\d{4}", mes_ano, b4s, flags=re.IGNORECASE)
            else:
                ws_model["B4"] = b4s if mes_ano in b4s else f"{b4s} - {mes_ano}".strip(" -")
        except Exception:
            pass

        ws_model["C8"] = responsavel.strip()
        ws_model["K8"] = now.strftime("%d/%m/%Y")

        # Verificação extra: turmas com 2+ profissionais distintos
        alerts = []
        def _turma_sort_key(x: str):
            # "2ºA" -> (2, "A")
            try:
                n = int(x.split("º")[0])
                l = x.split("º")[1]
                return (n, l)
            except Exception:
                return (999, x)

        for turma in sorted(valid_turmas, key=_turma_sort_key):
            prof_dict = profs_by_turma.get(turma, {})
            if len(prof_dict) >= 2:
                prof_names = sorted(
                    [prof_dict[k]["display"] for k in prof_dict.keys()],
                    key=lambda s: s.casefold(),
                )

                # auditoria opcional (leve): amostra por profissional
                audit = []
                for k in sorted(prof_dict.keys(), key=lambda s: prof_dict[s]["display"].casefold()):
                    alunos = prof_dict[k]["alunos"][:10]  # limite p/ UI
                    audit.append({
                        "profissional": prof_dict[k]["display"],
                        "amostra_alunos": [{"rm": rm, "nome": nome} for rm, nome in alunos],
                    })

                alerts.append({
                    "turma": turma,
                    "qtd_profissionais": len(prof_dict),
                    "profissionais": prof_names,
                    "auditoria": audit,
                })

        # Gera arquivo
        output = BytesIO()
        wb_model.save(output)
        output.seek(0)

        filename = f"Quadro_Quantitativo_de_Inclusao_{now.strftime('%d%m%Y')}.xlsx"
        resp = send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Envia alertas no header (ASCII-safe via base64)
        if alerts:
            payload = json.dumps(alerts, ensure_ascii=False)
            b64 = base64.b64encode(payload.encode("utf-8")).decode("ascii")
            if len(b64) > 7000:
                b64 = b64[:7000]
                resp.headers["X-QuantInclusao-Alerts-Truncated"] = "1"
            resp.headers["X-QuantInclusao-Alerts"] = b64

        return resp

    return render_template("quantinclusao.html")

import os
import json
import calendar
from dataclasses import dataclass
from datetime import datetime, date, timedelta

# ============================
# ALERTAS DE PRAZO (SEM DB)
# - usa feriados.json (dict: "YYYY-MM-DD": "Nome")
# - calcula D0 ajustado para dia útil conforme regra:
#   * Dia 20: próximo dia útil (posterior)
#   * Último dia do mês: último dia útil do mês (anterior, se não útil)
#   * Semanal: próximo dia útil (posterior) a partir do dia configurado
# - EXIBE SOMENTE NAS JANELAS PEDIDAS:
#   * Dia 20:        D-2, D-1, D0, D+1, D+2
#   * Último do mês: D-2, D-1, D0, D+1, D+2
#   * Semanal:       D-1, D0, D+1
# ============================

# Timezone (Render costuma rodar em UTC; aqui forçamos America/Sao_Paulo)
try:
    from zoneinfo import ZoneInfo
    _TZ = ZoneInfo("America/Sao_Paulo")
except Exception:
    _TZ = None


def _today_sp() -> date:
    if _TZ:
        return datetime.now(_TZ).date()
    return date.today()


# Caminho padrão do arquivo no projeto:
# .../secretariapadin/modelos/feriados.json
def _default_holidays_path() -> str:
    return os.path.join(app.root_path, "modelos", "feriados.json")


app.config.setdefault("HOLIDAYS_JSON_PATH", _default_holidays_path())

# Configurável: qual dia fecha o "Informativo Semanal"
# 0=Seg, 1=Ter, 2=Qua, 3=Qui, 4=Sex, 5=Sáb, 6=Dom
app.config.setdefault("INFORMATIVO_WEEKDAY_DUE", 4)  # padrão: sexta-feira


# ----------------------------
# Cache de feriados (memória)
# ----------------------------
_HOLIDAYS_CACHE = {
    "loaded": False,
    "dates": set(),      # set(date)
    "names": {},         # dict[date] = "Nome"
    "error": None,
    "path": None,
}


def _load_holidays_json_once() -> None:
    """
    Espera formato:
      {
        "2001-01-01": "Confraternização Universal",
        "2001-02-27": "Carnaval",
        ...
      }
    """
    if _HOLIDAYS_CACHE["loaded"]:
        return

    path = app.config.get("HOLIDAYS_JSON_PATH") or _default_holidays_path()
    _HOLIDAYS_CACHE["path"] = path

    try:
        if not os.path.exists(path):
            _HOLIDAYS_CACHE["loaded"] = True
            _HOLIDAYS_CACHE["error"] = f"feriados.json não encontrado em: {path}"
            return

        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)

        if not isinstance(data, dict):
            _HOLIDAYS_CACHE["loaded"] = True
            _HOLIDAYS_CACHE["error"] = "feriados.json inválido: esperado um objeto JSON (dict)."
            return

        for k, v in data.items():
            if not isinstance(k, str):
                continue
            # chave esperada: YYYY-MM-DD
            try:
                y = int(k[0:4]); m = int(k[5:7]); d = int(k[8:10])
                dt = date(y, m, d)
            except Exception:
                continue

            _HOLIDAYS_CACHE["dates"].add(dt)
            name = str(v).strip() if v is not None else ""
            if name:
                _HOLIDAYS_CACHE["names"][dt] = name

        _HOLIDAYS_CACHE["loaded"] = True
        _HOLIDAYS_CACHE["error"] = None

    except Exception as e:
        _HOLIDAYS_CACHE["loaded"] = True
        _HOLIDAYS_CACHE["error"] = str(e)


def _holiday_name(d: date) -> str:
    _load_holidays_json_once()
    return _HOLIDAYS_CACHE["names"].get(d, "")


def _is_business_day(d: date) -> bool:
    _load_holidays_json_once()
    if d.weekday() >= 5:  # 5=sábado, 6=domingo
        return False
    return d not in _HOLIDAYS_CACHE["dates"]


def _next_business_day(d: date) -> date:
    cur = d
    while not _is_business_day(cur):
        cur += timedelta(days=1)
    return cur


def _prev_business_day(d: date) -> date:
    cur = d
    while not _is_business_day(cur):
        cur -= timedelta(days=1)
    return cur


def _last_day_of_month(y: int, m: int) -> date:
    last = calendar.monthrange(y, m)[1]
    return date(y, m, last)


def _add_months(base: date, months: int) -> date:
    y = base.year + (base.month - 1 + months) // 12
    m = (base.month - 1 + months) % 12 + 1
    day = min(base.day, calendar.monthrange(y, m)[1])
    return date(y, m, day)


def _fmt_br(d: date) -> str:
    return d.strftime("%d/%m/%Y")


def _window_hit(today: date, due: date, before: int, after: int) -> bool:
    return (due - timedelta(days=before)) <= today <= (due + timedelta(days=after))


def _days_delta(today: date, due: date) -> int:
    return (due - today).days


@dataclass
class DeadlineAlert:
    key: str
    title: str
    due_base: date
    due_adjusted: date
    window_before: int
    window_after: int
    message: str


def _compute_due_day20(ref: date) -> tuple[date, date]:
    """
    D0 base: dia 20 do mês de referência.
    D0 ajustado: próximo dia útil (posterior).
    Se já passou da janela (D0 ajustado + 2), usa próximo mês.
    """
    base = date(ref.year, ref.month, 20)
    adj = _next_business_day(base)

    # janela: D-2..D+2
    if ref > (adj + timedelta(days=2)):
        base2 = _add_months(base, 1)
        base2 = date(base2.year, base2.month, 20)
        return base2, _next_business_day(base2)

    return base, adj


def _compute_due_month_end(ref: date) -> tuple[date, date]:
    """
    REGRA AJUSTADA CONFORME SEU ADENDO:

    D0 base: último dia do mês.
    D0 ajustado: ÚLTIMO DIA ÚTIL DO MÊS (se o último dia cair em não útil,
                 volta para o dia útil anterior).

    Se já passou da janela (D0 ajustado + 2), usa o mês seguinte.
    """
    base = _last_day_of_month(ref.year, ref.month)

    # Se o último dia não for útil, usa o último dia útil do mês (anterior)
    if _is_business_day(base):
        adj = base
    else:
        adj = _prev_business_day(base)

    # janela: D-2..D+2
    if ref > (adj + timedelta(days=2)):
        nextm = _add_months(date(ref.year, ref.month, 1), 1)
        base2 = _last_day_of_month(nextm.year, nextm.month)
        if _is_business_day(base2):
            adj2 = base2
        else:
            adj2 = _prev_business_day(base2)
        return base2, adj2

    return base, adj


def _compute_due_weekly(ref: date) -> tuple[date, date]:
    """
    D0 base: dia da semana configurado (default sexta) na semana de ref.
    D0 ajustado: próximo dia útil (posterior) se cair em não útil/feriado.
    Se já passou da janela (D0 ajustado + 1), usa próxima semana.
    """
    due_wd = int(app.config.get("INFORMATIVO_WEEKDAY_DUE", 4))
    due_wd = max(0, min(6, due_wd))

    # segunda-feira da semana atual
    monday = ref - timedelta(days=ref.weekday())
    base = monday + timedelta(days=due_wd)
    adj = _next_business_day(base)

    # janela semanal: D-1..D+1
    if ref > (adj + timedelta(days=1)):
        monday2 = monday + timedelta(days=7)
        base2 = monday2 + timedelta(days=due_wd)
        return base2, _next_business_day(base2)

    return base, adj


def build_deadline_alerts(today: date | None = None) -> list[dict]:
    """
    Retorna lista de dicts pronta para template.
    Exibe somente quando estiver dentro da janela correta.
    """
    today = today or _today_sp()

    # monta D0 base e ajustado por quadro
    day20_base, day20_adj = _compute_due_day20(today)
    mend_base, mend_adj = _compute_due_month_end(today)
    wk_base, wk_adj = _compute_due_weekly(today)

    alerts: list[DeadlineAlert] = []

    # 1) Quantitativo de Inclusão (dia 20 ou próximo útil) -> D-2..D+2
    if _window_hit(today, day20_adj, before=2, after=2):
        extra = ""
        if day20_adj != day20_base:
            hn = _holiday_name(day20_base)
            extra = f" (ajustado para próximo dia útil{': ' + hn if hn else ''})"
        alerts.append(DeadlineAlert(
            key="quant_inclusao",
            title="Prazo: Quantitativo de Inclusão",
            due_base=day20_base,
            due_adjusted=day20_adj,
            window_before=2,
            window_after=2,
            message=f"O Quadro Quantitativo de Inclusão deve ser enviado até {_fmt_br(day20_adj)}{extra}."
        ))

    # 2) Atendimento Mensal (último dia útil do mês) -> D-2..D+2
    if _window_hit(today, mend_adj, before=2, after=2):
        extra = ""
        if mend_adj != mend_base:
            hn = _holiday_name(mend_base)
            extra = f" (ajustado para último dia útil do mês{': ' + hn if hn else ''})"
        alerts.append(DeadlineAlert(
            key="atendimento_mensal",
            title="Prazo: Atendimento Mensal",
            due_base=mend_base,
            due_adjusted=mend_adj,
            window_before=2,
            window_after=2,
            message=f"O Quadro de Atendimento Mensal deve ser enviado até {_fmt_br(mend_adj)}{extra}."
        ))

    # 3) Quantitativo Mensal de Transferências Expedidas (último dia útil do mês) -> D-2..D+2
    if _window_hit(today, mend_adj, before=2, after=2):
        extra = ""
        if mend_adj != mend_base:
            hn = _holiday_name(mend_base)
            extra = f" (ajustado para último dia útil do mês{': ' + hn if hn else ''})"
        alerts.append(DeadlineAlert(
            key="quant_mensal_te",
            title="Prazo: Quantitativo Mensal de Transferências Expedidas",
            due_base=mend_base,
            due_adjusted=mend_adj,
            window_before=2,
            window_after=2,
            message=f"O Quadro Mensal de Transferências Expedidas deve ser enviado até {_fmt_br(mend_adj)}{extra}."
        ))

    # 4) Informativo Semanal -> D-1..D+1
    if _window_hit(today, wk_adj, before=1, after=3):
        extra = ""
        if wk_adj != wk_base:
            hn = _holiday_name(wk_base)
            extra = f" (ajustado para próximo dia útil{': ' + hn if hn else ''})"
        alerts.append(DeadlineAlert(
            key="informativo_semanal",
            title="Prazo: Informativo Semanal",
            due_base=wk_base,
            due_adjusted=wk_adj,
            window_before=1,
            window_after=1,
            message=f"O Informativo Semanal deve ser enviado até {_fmt_br(wk_adj)}{extra}."
        ))

    # converte para dict simples + status por D- / D+
    out: list[dict] = []
    for a in alerts:
        delta = _days_delta(today, a.due_adjusted)
        if delta > 0:
            when = f"Faltam {delta} dia(s)."
        elif delta == 0:
            when = "Vence hoje."
        else:
            when = f"Vencido há {abs(delta)} dia(s)."

        out.append({
            "key": a.key,
            "title": a.title,
            "message": a.message,
            "due": _fmt_br(a.due_adjusted),
            "status": when,
        })

    return out


# Injeta automaticamente para TODOS os templates
@app.context_processor
def _inject_deadline_alerts():
    return {"deadline_alerts": build_deadline_alerts()}

# ==========================================================
#  MAIN
# ==========================================================

if __name__ == "__main__":
    app.run(debug=True)
